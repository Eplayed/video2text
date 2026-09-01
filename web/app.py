"""video2text Web 管理界面 - Flask Backend"""
import json, os, sys, sqlite3, threading, time, traceback, hashlib
from pathlib import Path
from datetime import datetime

from flask import Flask, jsonify, request, send_from_directory

# ── 添加项目根到 path ──
ROOT = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(ROOT))

import main as collector  # main.py
from src import content_store, material_store
from src.dify_client import DifyKBClient, DifyKBError
from src.path_config import find_parser_dir, ensure_parser_on_path, get_cookie_path

# ── 路径 ──
OUTPUT_DIR    = ROOT / "output"
EXCEL_PATH    = OUTPUT_DIR / "抖音视频信息.xlsx"
INDEX_PATH    = ROOT / "video_index.json"
DB_PATH       = OUTPUT_DIR / "video2text.db"
PARSER_DIR    = Path(find_parser_dir())

app = Flask(__name__, static_folder="static", static_url_path="/static")

# ── 处理状态（内存） ──
_task_status = {"running": False, "progress": "", "done": False, "error": ""}
_classify_status = {"running": False, "progress": "", "done": True, "error": "", "result": None}
_dify_status = {"running": False, "progress": "", "done": True, "error": "", "result": None}
_db_lock = threading.RLock()

# ── 工具：cookie 写入与 parser 初始化 ──
def _setup_parser_and_cookie(cookie_str: str):
    """写入 cookie 文件并初始化 parser（自动加上 sessionid= 前缀）。
    不使用 os.chdir（会影响 Flask 工作目录），而是通过 set_cookie() 直接设置。
    """
    # 自动加上 sessionid= 前缀
    if cookie_str and "sessionid=" not in cookie_str:
        cookie_str = f"sessionid={cookie_str}"

    parser_dir = str(PARSER_DIR)
    if not os.path.isdir(parser_dir):
        raise RuntimeError(
            f"douyin_parse 目录不存在: {parser_dir}\n"
            f"请执行: git clone https://github.com/DLWangSan/douyin_parse.git {parser_dir}"
        )

    # 写入 cookie 文件（部分内部逻辑仍可能读取文件，保持兼容）
    cookie_path = get_cookie_path()
    cookie_path.parent.mkdir(parents=True, exist_ok=True)
    with open(cookie_path, "w") as f:
        f.write(cookie_str)

    # 确保 parser 目录在 sys.path 中并创建实例
    ensure_parser_on_path()
    parser = collector.get_douyin_parser(parser_dir)
    # 直接设置 cookie，避免依赖 os.chdir
    parser.set_cookie(cookie_str)
    return parser


def _sync_content_db():
    """把 Excel 当前内容同步到 SQLite 内容库。"""
    if not EXCEL_PATH.exists():
        raise RuntimeError(f"Excel 不存在: {EXCEL_PATH}")
    DB_PATH.parent.mkdir(parents=True, exist_ok=True)
    with _db_lock:
        return content_store.sync_excel_to_db(EXCEL_PATH, DB_PATH)


def _sync_material_workspace():
    """同步 SQLite + AI JSONL 素材层，轻量操作，不下载媒体。"""
    with _db_lock:
        content_store.sync_excel_to_db(EXCEL_PATH, DB_PATH)
        return material_store.export_workspace(EXCEL_PATH, DB_PATH)


def _ai_config():
    config = collector.load_env_file(str(ROOT / "config" / "config.env.local"))
    method = collector.normalize_ai_method(collector.config_get(config, "AI_METHOD", "skip"))
    if method == "deepseek":
        api_key = collector.config_get(config, "DEEPSEEK_API_KEY") or collector.config_get(config, "AI_API_KEY")
        api_base = collector.config_get(config, "DEEPSEEK_API_BASE", "https://api.deepseek.com")
        model = collector.config_get(config, "DEEPSEEK_MODEL", "deepseek-chat")
    else:
        api_key = collector.config_get(config, "OPENAI_API_KEY") or collector.config_get(config, "AI_API_KEY")
        api_base = collector.config_get(config, "AI_API_BASE", "https://api.openai.com/v1")
        model = collector.config_get(config, "AI_MODEL", "gpt-4o-mini")
    return {"method": method, "api_key": api_key, "api_base": api_base, "model": model}

# ── API: 视频列表 ──
@app.route("/api/videos")
def api_videos():
    topic = request.args.get("topic", "")
    q     = request.args.get("q", "").lower()
    category = request.args.get("category", "")
    game = request.args.get("game", "")
    if not INDEX_PATH.exists():
        return jsonify({"videos": [], "total": 0})
    with open(INDEX_PATH, encoding="utf-8") as f:
        data = json.load(f)
    videos = data.get("videos", [])
    # 合并 SQLite 中的分类/标签/Dify 同步状态（按 sheet+row 关联）
    db_info = _load_video_extras()
    for v in videos:
        info = db_info.get((v.get("sheet"), v.get("row")))
        if info:
            v["db_id"] = info["id"]
            v["category"] = info["category"]
            v["ai_tags"] = info["ai_tags"]
            v["game"] = info["game"]
            v["dify_synced_at"] = info["dify_synced_at"]
        else:
            v.setdefault("category", "")
            v.setdefault("ai_tags", "")
            v.setdefault("game", "")
            v.setdefault("dify_synced_at", "")

    # 追加 SQLite 中的微信文章（不经过 video_index.json）
    if DB_PATH.exists():
        try:
            with _db_lock:
                conn = content_store.connect(DB_PATH)
                try:
                    wx_rows = conn.execute(
                        "SELECT id, source_sheet, source_row, source_url, author, title, "
                        "status, published_at, category, ai_tags, game, dify_synced_at, "
                        "COALESCE(cover_url,'') as cover "
                        "FROM videos WHERE source_sheet = '微信文章' ORDER BY id DESC LIMIT 200"
                    ).fetchall()
                finally:
                    conn.close()
            for r in wx_rows:
                videos.append({
                    "id": r["id"], "db_id": r["id"],
                    "sheet": r["source_sheet"], "row": r["source_row"],
                    "url": r["source_url"] or "",
                    "aweme_id": "", "author": r["author"] or "",
                    "title": r["title"] or "", "status": r["status"] or "",
                    "published_at": r["published_at"] or "",
                    "cover": r["cover"] or "",
                    "category": r["category"] or "",
                    "ai_tags": r["ai_tags"] or "",
                    "game": r["game"] or "",
                    "dify_synced_at": r["dify_synced_at"] or "",
                })
        except Exception:
            pass
    if topic:
        videos = [v for v in videos if v.get("author", "") == topic]
    if category:
        want = category if category != "未分类" else ""
        videos = [v for v in videos if (v.get("category") or "") == want]
    if game:
        want_game = "" if game == "未识别游戏" else game
        videos = [v for v in videos if (v.get("game") or "") == want_game]
    if q:
        # 按空白分词、逐词 AND 匹配（标题/作者/描述/分类/标签/游戏/工作表/行号）。
        # 整串关键词或粘贴整行卡片（如"…送英雄装 游戏攻略 泰莉亚子 抖音视频数据 R404"）都能命中；
        # 纯符号/emoji 词直接忽略，R404 / 404 均可命中行号。
        import re as _re
        tokens = [t.lower() for t in q.split() if _re.search(r"\w", t)]
        if tokens:
            def _haystack(v):
                return " ".join([
                    v.get("title", ""), v.get("author", ""), v.get("description", ""),
                    v.get("category", ""), v.get("game", ""), v.get("ai_tags", ""),
                    v.get("sheet", ""), f"R{v.get('row', '')}", str(v.get("row", "")),
                ]).lower()
            videos = [v for v in videos if all(t in _haystack(v) for t in tokens)]
    # 按时间最新在前（发布时间优先，其次创建时间；空值垫底）
    videos.sort(
        key=lambda v: (v.get("published_at") or v.get("create_time") or v.get("pub_time") or ""),
        reverse=True,
    )
    return jsonify({"videos": videos, "total": len(videos)})


def _load_video_extras() -> dict:
    """从 SQLite 取分类/Dify 字段，键为 (sheet, row)。"""
    if not DB_PATH.exists():
        return {}
    try:
        with _db_lock:
            conn = content_store.connect(DB_PATH)
            try:
                rows = conn.execute(
                    "SELECT id, source_sheet, source_row, category, ai_tags, game, dify_synced_at FROM videos"
                ).fetchall()
            finally:
                conn.close()
        return {
            (r["source_sheet"], r["source_row"]): {
                "id": r["id"],
                "category": r["category"] or "",
                "ai_tags": r["ai_tags"] or "",
                "game": r["game"] or "",
                "dify_synced_at": r["dify_synced_at"] or "",
            }
            for r in rows
        }
    except Exception:
        return {}


# ── API: 分类统计 ──
@app.route("/api/categories")
def api_categories():
    if not DB_PATH.exists():
        return jsonify({"categories": [], "games": []})
    try:
        with _db_lock:
            stats = content_store.get_category_stats(DB_PATH)
            games = content_store.get_game_stats(DB_PATH)
    except Exception as e:
        return jsonify({"categories": [], "games": [], "error": str(e)}), 500
    return jsonify({"categories": stats, "games": games})


# ── API: 选题雷达 ──
@app.route("/api/topics/radar")
def api_topic_radar():
    """选题雷达：标签热度聚合。category/game/author 可选筛选，channel 切换渠道策略。"""
    category = (request.args.get("category") or "").strip()
    game = (request.args.get("game") or "").strip()
    author = (request.args.get("author") or "").strip()
    channel = (request.args.get("channel") or "").strip()
    if channel and channel not in content_store.RADAR_CHANNEL_STRATEGY:
        channel = ""
    if not DB_PATH.exists():
        return jsonify({"topics": [], "total_videos": 0, "window_days": 90})
    try:
        with _db_lock:
            data = content_store.get_topic_radar(
                DB_PATH, category, game, author, channel=channel
            )
        return jsonify(data)
    except Exception as e:
        return jsonify({"error": str(e)}), 500


# ── API: 渠道策略配置下发（自媒体工作台策略引擎的单一权威源） ──
@app.route("/api/strategy/channels")
def api_strategy_channels():
    """工作台前端策略引擎（channel-strategy.js）从这里拉取策略口径。"""
    return jsonify(content_store.get_channel_strategy_config())


# ── API: AI 批量分类 ──
def _auto_classify_after_sync() -> dict:
    """同步入库后的增量自动打标：只补 category/ai_tags 任一为空的视频。

    - AI 未配置（method=skip 或无 key）时静默返回空结果，不阻塞采集主流程；
    - 沿用 classify_videos 的 LLM 优先、失败降级规则分类；
    - 调用方在采集线程里，此函数同步执行（单条链接量小可接受）。
      持锁口径与 /api/videos/classify 一致。
    """
    try:
        cfg = _ai_config()
        if cfg.get("method") == "skip" or not cfg.get("api_key"):
            return {"total": 0, "classified": 0, "llm_batches": 0, "skipped": "ai_not_configured"}
        with _db_lock:
            return content_store.classify_videos(DB_PATH, cfg, force=False)
    except Exception as e:
        # 打标失败不算采集失败：记日志，返回空结果让主流程继续
        print(f"[auto-classify] 增量打标失败（不影响采集）: {e}", flush=True)
        return {"total": 0, "classified": 0, "llm_batches": 0, "error": str(e)}


@app.route("/api/videos/classify", methods=["POST"])
def api_classify_videos():
    if _classify_status.get("running"):
        return jsonify({"error": "分类任务进行中，请稍候"}), 409
    force = bool((request.get_json(force=True, silent=True) or {}).get("force"))

    def run():
        _classify_status.update(running=True, done=False, error="", progress="开始分类...", result=None)
        try:
            def cb(done, total, note):
                _classify_status["progress"] = note
            with _db_lock:
                result = content_store.classify_videos(DB_PATH, _ai_config(), force=force, progress_cb=cb)
            _classify_status["result"] = result
            _classify_status["progress"] = f"完成：{result['classified']}/{result['total']} 条已分类"
        except Exception as e:
            _classify_status["error"] = f"{e}\n{traceback.format_exc()[-500:]}"
        finally:
            _classify_status["running"] = False
            _classify_status["done"] = True

    threading.Thread(target=run, daemon=True).start()
    return jsonify({"success": True})


@app.route("/api/classify/status")
def api_classify_status():
    return jsonify(_classify_status)


# ── API: 手动修改分类 ──
@app.route("/api/videos/category", methods=["POST"])
def api_update_video_category():
    data = request.get_json(force=True)
    sheet = (data.get("sheet") or "").strip()
    row = data.get("row")
    category = (data.get("category") or "").strip()
    game = data.get("game")
    if not sheet or not row:
        return jsonify({"error": "缺少 sheet/row"}), 400
    try:
        with _db_lock:
            conn = content_store.connect(DB_PATH)
            try:
                r = conn.execute(
                    "SELECT id FROM videos WHERE source_sheet = ? AND source_row = ?", (sheet, row)
                ).fetchone()
            finally:
                conn.close()
        if not r:
            return jsonify({"error": "视频不存在于内容库，请先同步内容库"}), 404
        content_store.update_video_category(DB_PATH, r["id"], category, game=game)
        return jsonify({"success": True})
    except Exception as e:
        return jsonify({"error": str(e)}), 500

# ── API: 视频详情 ──
@app.route("/api/videos/<sheet>/<int:row>")
def api_video_detail(sheet, row):
    import openpyxl
    if not EXCEL_PATH.exists():
        return jsonify({"error": "Excel 不存在"}), 404
    wb = openpyxl.load_workbook(str(EXCEL_PATH), read_only=True)
    try:
        ws = wb[sheet]
    except KeyError:
        return jsonify({"error": f"Sheet '{sheet}' 不存在"}), 404
    if row < 2 or row > ws.max_row:
        return jsonify({"error": f"行号 {row} 超出范围"}), 404
    cols = {"link":1,"status":2,"id":3,"author":4,"pub_time":5,"title":6,
            "asr":7,"tags":8,"cover":9,"video_url":10,"raw_transcript":11,
            "ai_copy":12,"ai_title":13,"keywords":14,"remark":15}
    detail = {k: str(ws.cell(row, c).value or "") for k, c in cols.items()}
    detail["sheet"] = sheet
    detail["row"] = row
    wb.close()
    # 已抽取的关键帧（生成攻略文章时落盘），供自媒体工作台取材配图
    try:
        aweme_id = material_store.extract_aweme_id(detail.get("link") or detail.get("video_url") or "")
        frame_paths = [p for p in material_store.local_keyframe_paths(aweme_id) if not p.endswith("sheet.jpg")]
        detail["keyframes"] = [f"/media/{p}" for p in frame_paths]
    except Exception:
        detail["keyframes"] = []
    # 该视频最新的公众号素材档案（wechat_material 整理稿），供工作台取材时优先于原始转写
    try:
        if DB_PATH.exists():
            video = content_store.get_video_by_source(DB_PATH, sheet, row)
            if video:
                with _db_lock:
                    conn = content_store.connect(DB_PATH)
                    summary_row = conn.execute(
                        "SELECT id, title, content FROM ai_summaries "
                        "WHERE video_id = ? AND summary_type = 'wechat_material' "
                        "ORDER BY id DESC LIMIT 1",
                        (video["id"],),
                    ).fetchone()
                    conn.close()
                if summary_row:
                    detail["wechat_material"] = summary_row["content"] or ""
                    detail["wechat_material_title"] = summary_row["title"] or ""
    except Exception:
        detail["wechat_material"] = ""
    return jsonify(detail)


@app.route("/media/<path:filename>")
def api_media(filename):
    """Serve local output assets such as covers and keyframes."""
    return send_from_directory(OUTPUT_DIR, filename)


@app.route("/api/workbench")
def api_workbench():
    try:
        if not DB_PATH.exists():
            _sync_content_db()
        q = request.args.get("q", "")
        author = request.args.get("author", "")
        status = request.args.get("status", "")
        with _db_lock:
            items = material_store.list_workbench(DB_PATH, query=q, author=author, status=status)
        return jsonify({"items": items, "total": len(items)})
    except Exception as e:
        return jsonify({"items": [], "total": 0, "error": str(e)}), 500


@app.route("/api/workbench/sync", methods=["POST"])
def api_workbench_sync():
    try:
        result = _sync_material_workspace()
        return jsonify({"success": True, **result})
    except Exception as e:
        return jsonify({"success": False, "error": str(e)}), 500

# ── API: 处理状态 ──
@app.route("/api/status")
def api_status():
    return jsonify(_task_status)

# ── API: 统计 ──
@app.route("/api/stats")
def api_stats():
    if not INDEX_PATH.exists():
        return jsonify({"topics": {}, "total": 0})
    with open(INDEX_PATH, encoding="utf-8") as f:
        data = json.load(f)
    topics = {}
    for v in data.get("videos", []):
        author = v.get("author", "未知作者") or "未知作者"
        topics[author] = topics.get(author, 0) + 1
    return jsonify({"topics": topics, "total": data.get("total", 0),
                     "version": data.get("version")})


# ── API: 工作台仪表盘 ──
@app.route("/api/dashboard")
def api_dashboard():
    """聚合 SQLite 内容库统计，供首页仪表盘使用。"""
    try:
        if not DB_PATH.exists():
            _sync_content_db()
        with _db_lock:
            conn = sqlite3.connect(str(DB_PATH))
            conn.row_factory = sqlite3.Row
            try:
                total = conn.execute("SELECT COUNT(*) AS c FROM videos").fetchone()["c"]
                with_asr = conn.execute(
                    "SELECT COUNT(*) AS c FROM videos WHERE transcript IS NOT NULL AND transcript != ''"
                ).fetchone()["c"]
                authors = conn.execute(
                    "SELECT author AS name, COUNT(*) AS c FROM videos "
                    "GROUP BY author ORDER BY c DESC LIMIT 8"
                ).fetchall()
                summary_rows = conn.execute(
                    "SELECT summary_type, COUNT(*) AS c FROM ai_summaries GROUP BY summary_type"
                ).fetchall()
                recent_videos = conn.execute(
                    "SELECT title, author, status, updated_at FROM videos "
                    "ORDER BY updated_at DESC, id DESC LIMIT 6"
                ).fetchall()
                recent_summaries = conn.execute(
                    "SELECT id, title, summary_type, status, updated_at FROM ai_summaries "
                    "ORDER BY updated_at DESC, id DESC LIMIT 6"
                ).fetchall()
                question_count = 0
                for row in conn.execute(
                    "SELECT structured_data FROM ai_summaries WHERE summary_type = 'ai_interview'"
                ).fetchall():
                    try:
                        payload = json.loads(row["structured_data"] or "{}")
                        question_count += len(payload.get("questions") or [])
                    except Exception:
                        pass
            finally:
                conn.close()

        summaries = {row["summary_type"]: row["c"] for row in summary_rows}
        return jsonify({
            "total": total,
            "with_asr": with_asr,
            "missing_asr": max(0, total - with_asr),
            "asr_rate": round(with_asr * 100 / total, 1) if total else 0,
            "authors": [{"name": r["name"] or "未知作者", "count": r["c"]} for r in authors],
            "summaries": {
                "game_guide": summaries.get("game_guide", 0),
                "ai_interview": summaries.get("ai_interview", 0),
                "wechat_material": summaries.get("wechat_material", 0),
            },
            "question_count": question_count,
            "recent_videos": [dict(r) for r in recent_videos],
            "recent_summaries": [dict(r) for r in recent_summaries],
            "running": _task_status.get("running", False),
        })
    except Exception as e:
        return jsonify({"error": str(e)}), 500

# ── API: 提交处理 ──
@app.route("/api/process", methods=["POST"])
def api_process():
    global _task_status
    if _task_status["running"]:
        return jsonify({"error": "已有任务正在运行"}), 400

    data = request.get_json(force=True)
    links_text = data.get("links", "").strip()
    cookie = data.get("cookie", "").strip()
    asr_model = data.get("asr_model", "base")

    if not links_text:
        return jsonify({"error": "请输入抖音链接"}), 400
    if not cookie:
        return jsonify({"error": "请输入 Cookie"}), 400

    _task_status = {"running": True, "progress": "准备中...", "done": False, "error": ""}

    def run():
        global _task_status
        import openpyxl
        try:
            # 1. 初始化 parser
            _task_status["progress"] = "初始化解析器..."
            parser = _setup_parser_and_cookie(cookie)

            # 2. 写入链接到 Excel
            _task_status["progress"] = "写入链接到 Excel..."
            wb = openpyxl.load_workbook(str(EXCEL_PATH))
            ws = wb["抖音视频数据"]
            next_row = ws.max_row + 1
            # 找第一个空行
            for r in range(2, ws.max_row + 1):
                if not ws.cell(r, 1).value:
                    next_row = r
                    break
            link_list = [l.strip() for l in links_text.split("\n") if l.strip()]
            for i, link in enumerate(link_list):
                ws.cell(int(next_row) + i, 1).value = link
                ws.cell(int(next_row) + i, 2).value = "未开始"
            wb.save(str(EXCEL_PATH))
            wb.close()

            # 3. 处理每条链接（重新打开 Excel 让 process_row 操作）
            ai_config = {"method": "skip", "api_key": "", "api_base": "", "model": ""}

            for i, link in enumerate(link_list):
                row = int(next_row) + i
                _task_status["progress"] = f"处理 [{i+1}/{len(link_list)}] 第 {row} 行..."

                wb = openpyxl.load_workbook(str(EXCEL_PATH))
                ws = wb["抖音视频数据"]
                ok = collector.process_row(ws, row, cookie, asr_model, ai_config, parser)
                wb.save(str(EXCEL_PATH))
                wb.close()

                if not ok:
                    _task_status["error"] += f"Row {row}: 处理失败\n"

                if i < len(link_list) - 1:
                    import time
                    time.sleep(5)

            # 4. 更新索引 + 增量自动打标
            _task_status["progress"] = "更新索引..."
            collector.update_video_index(str(EXCEL_PATH))
            _sync_content_db()

            _task_status["progress"] = "AI 自动打标..."
            _cls = _auto_classify_after_sync()
            if _cls.get("classified"):
                _task_status["progress"] = f"AI 打标 {_cls['classified']}/{_cls['total']} 条完成"

            _task_status["progress"] = "全部处理完成 ✅"
            _task_status["done"] = True

        except Exception as e:
            _task_status["error"] = traceback.format_exc()
            _task_status["done"] = True
        finally:
            _task_status["running"] = False

    threading.Thread(target=run, daemon=True).start()
    return jsonify({"status": "started"})

# ── 前端页面 ──
# ── API: 批量获取用户视频列表 ──────────────────────────────────────
@app.route("/api/fetch_user_videos", methods=["POST"])
def api_fetch_user_videos():
    data = request.get_json(force=True)
    url = (data.get("user_url") or data.get("video_url") or "").strip()
    cookie = data.get("cookie", "").strip()
    max_pages = int(data.get("max_pages", 10))
    max_videos = int(data.get("max_videos") or 0)

    if not url:
        return jsonify({"error": "请输入用户主页链接或视频链接"}), 400
    if not cookie:
        return jsonify({"error": "请输入 Cookie"}), 400

    try:
        from src.fetch_user_videos import fetch_user_videos
        mode = "user_url" if "/user/" in url else "video_url"
        result = fetch_user_videos(
            url=url, cookie=cookie,
            max_pages=max_pages, max_videos=max_videos, mode=mode,
            exclude_excel=str(EXCEL_PATH),
        )
        return jsonify(result)
    except Exception as e:
        return jsonify({"error": str(e), "success": False}), 500


# ── API: 预览用户信息（不写入Excel，只获取列表） ────────────────────
@app.route("/api/preview_user_videos", methods=["POST"])
def api_preview_user_videos():
    """获取用户视频列表预览，不写入 Excel"""
    data = request.get_json(force=True)
    url = (data.get("user_url") or data.get("video_url") or "").strip()
    cookie = data.get("cookie", "").strip()
    max_videos = int(data.get("max_videos") or 0)
    max_pages = int(data.get("max_pages", 1))
    if max_videos <= 0:
        max_pages = min(max_pages, 5)

    # DEBUG LOG
    import json as _j
    print(f"[DEBUG] preview_user_videos: url={url[:50]}... cookie={cookie[:20]}... max_pages={max_pages} max_videos={max_videos}", flush=True)

    if not url:
        return jsonify({"error": "请输入链接"}), 400
    if not cookie:
        return jsonify({"error": "请输入 Cookie"}), 400

    try:
        from src.fetch_user_videos import fetch_user_videos
        mode = "user_url" if "/user/" in url else "video_url"
        result = fetch_user_videos(
            url=url, cookie=cookie,
            max_pages=max_pages, max_videos=max_videos, mode=mode,
            exclude_excel=str(EXCEL_PATH),
        )
        # 如果 total=0，给出提示（但如果是去重导致的，不提示）
        if result.get('total', 0) == 0 and not result.get('error'):
            # 检查是否是去重导致的
            if result.get('filtered'):
                result['warning'] = f"所有视频已存在（去重 {result.get('filtered')} 条）"
            else:
                result['warning'] = "视频数为0，请检查Cookie是否有效或已过期"
        return jsonify(result)
    except Exception as e:
        import traceback
        traceback.print_exc()
        return jsonify({"error": str(e), "success": False}), 500


# ── API: 批量获取 + 写入 Excel + 自动处理（全自动流水线） ─────────
@app.route("/api/fetch_and_process", methods=["POST"])
def api_fetch_and_process():
    """
    完整流水线：
    1. 获取用户主页视频列表
    2. 写入 Excel（状态=未开始）
    3. 后台逐条 ASR 处理
    4. 更新索引
    """
    global _task_status
    if _task_status["running"]:
        return jsonify({"error": "已有任务正在运行，请等待完成"}), 400

    data = request.get_json(force=True)
    url = (data.get("user_url") or data.get("video_url") or "").strip()
    cookie = data.get("cookie", "").strip()
    max_pages = int(data.get("max_pages", 10))
    max_videos = int(data.get("max_videos") or 0)
    asr_model = data.get("asr_model", "base")
    # 前端预览阶段已抓取的视频列表（可选，传入则跳过重复抓取）
    preset_videos = data.get("videos") or []

    if not url and not preset_videos:
        return jsonify({"error": "请输入用户主页链接或视频链接"}), 400
    if not cookie:
        return jsonify({"error": "请输入 Cookie"}), 400

    _task_status = {"running": True, "progress": "准备中...", "done": False, "error": "", "total": 0, "current": 0}

    def run():
        global _task_status
        import openpyxl, time
        try:
            # 1. 获取视频列表（优先使用前端预览结果）
            if preset_videos:
                videos = preset_videos
                _task_status["progress"] = f"使用预览列表 {len(videos)} 条..."
            else:
                _task_status["progress"] = "获取视频列表..."
                from src.fetch_user_videos import fetch_user_videos
                mode = "user_url" if "/user/" in url else "video_url"
                result = fetch_user_videos(url=url, cookie=cookie, max_pages=max_pages, max_videos=max_videos, mode=mode,
                                   exclude_excel=str(EXCEL_PATH))
                if not result.get("success") or result.get("total", 0) == 0:
                    _task_status["error"] = f"获取视频列表失败: {result.get('error', '0条视频')}"
                    _task_status["done"] = True
                    _task_status["running"] = False
                    return
                videos = result["videos"]
            _task_status["total"] = len(videos)
            _task_status["progress"] = f"写入 {len(videos)} 条视频到 Excel..."

            # 2. 写入 Excel
            wb = openpyxl.load_workbook(str(EXCEL_PATH))
            ws = wb["抖音视频数据"]
            next_row = ws.max_row + 1
            for r in range(2, ws.max_row + 1):
                if not ws.cell(r, 1).value:
                    next_row = r
                    break

            for i, v in enumerate(videos):
                ws.cell(next_row + i, 1).value = v["url"]
                ws.cell(next_row + i, 2).value = "未开始"
                ws.cell(next_row + i, 3).value = v.get("aweme_id", "")

            wb.save(str(EXCEL_PATH))
            wb.close()

            # 3. 初始化 parser
            _task_status["progress"] = "初始化解析器..."
            parser = _setup_parser_and_cookie(cookie)
            ai_config = {"method": "skip", "api_key": "", "api_base": "", "model": ""}

            # 4. 逐条 ASR 处理
            for i, v in enumerate(videos):
                row = next_row + i
                _task_status["current"] = i + 1
                _task_status["progress"] = f"处理 [{i+1}/{len(videos)}] {v.get('aweme_id','')}..."

                wb = openpyxl.load_workbook(str(EXCEL_PATH))
                ws = wb["抖音视频数据"]
                ok = collector.process_row(ws, row, cookie, asr_model, ai_config, parser)
                wb.save(str(EXCEL_PATH))
                wb.close()

                if not ok:
                    _task_status["error"] += f"Row {row}: 处理失败\n"

                if i < len(videos) - 1:
                    time.sleep(5)

            # 5. 更新索引 + 增量自动打标
            _task_status["progress"] = "更新索引..."
            collector.update_video_index(str(EXCEL_PATH))
            _sync_content_db()

            _task_status["progress"] = "AI 自动打标..."
            _cls = _auto_classify_after_sync()
            if _cls.get("classified"):
                _task_status["progress"] = f"AI 打标 {_cls['classified']}/{_cls['total']} 条完成"

            _task_status["progress"] = f"✅ 全部 {len(videos)} 条处理完成！"
            _task_status["done"] = True

        except Exception as e:
            _task_status["error"] = traceback.format_exc()
            _task_status["done"] = True
        finally:
            _task_status["running"] = False

    threading.Thread(target=run, daemon=True).start()
    return jsonify({"status": "started"})


# ── API: SQLite 内容库 ────────────────────────────────────────────
@app.route("/api/content/sync", methods=["POST"])
def api_content_sync():
    try:
        result = _sync_content_db()
        return jsonify({"success": True, **result, "db_path": str(DB_PATH)})
    except Exception as e:
        return jsonify({"success": False, "error": str(e)}), 500


@app.route("/api/content/summaries")
def api_content_summaries():
    try:
        if not DB_PATH.exists():
            _sync_content_db()
        summary_type = request.args.get("type", "")
        q = request.args.get("q", "")
        items = content_store.list_summaries(DB_PATH, summary_type=summary_type, query=q)
        return jsonify({"items": items, "total": len(items)})
    except Exception as e:
        return jsonify({"error": str(e), "items": [], "total": 0}), 500


@app.route("/api/content/summaries/<int:summary_id>")
def api_content_summary_detail(summary_id):
    try:
        item = content_store.get_summary(DB_PATH, summary_id)
        if not item:
            return jsonify({"error": "总结不存在"}), 404
        return jsonify(item)
    except Exception as e:
        return jsonify({"error": str(e)}), 500


@app.route("/api/content/summaries/<int:summary_id>/regenerate", methods=["POST"])
def api_content_regenerate(summary_id):
    """重新生成指定整理稿。"""
    try:
        summary = content_store.get_summary(DB_PATH, summary_id)
        if not summary:
            return jsonify({"error": "整理稿不存在"}), 404

        summary_type = request.args.get("type") or summary.get("summary_type", "game_guide")
        config = _ai_config()
        source_ids = json.loads(summary.get("source_video_ids") or "[]")
        if len(source_ids) > 1:
            videos = content_store.get_videos_by_ids(DB_PATH, [int(v) for v in source_ids])
            if not videos:
                return jsonify({"error": "关联视频不存在"}), 404
            result, model, status = content_store.generate_collection_summary(
                videos, summary_type, config
            )
            content_store.update_summary(
                DB_PATH, summary_id, result, model, status, source_video_ids=[int(v["id"]) for v in videos]
            )
        else:
            video = content_store.get_video_by_source(DB_PATH, summary["source_sheet"], summary["source_row"])
            if not video:
                return jsonify({"error": "关联视频不存在"}), 404
            result, model, status = content_store.generate_summary(video, summary_type, config)
            content_store.update_summary(
                DB_PATH, summary_id, result, model, status, source_video_ids=[int(video["id"])]
            )

        updated = content_store.get_summary(DB_PATH, summary_id)
        return jsonify({"success": True, "summary": updated})
    except Exception as e:
        import traceback; traceback.print_exc()
        return jsonify({"error": str(e)}), 500



@app.route("/api/content/summaries/<int:summary_id>", methods=["DELETE"])
def api_content_delete(summary_id):
    """删除指定整理稿"""
    try:
        content_store.delete_summary(DB_PATH, summary_id)
        return jsonify({"success": True, "deleted": summary_id})
    except Exception as e:
        return jsonify({"error": str(e)}), 500


@app.route("/api/content/generate", methods=["POST"])
def api_content_generate():
    """把选中的视频整理成游戏攻略或 AI 面试题，并保存到 SQLite。"""
    global _task_status
    if _task_status.get("running"):
        return jsonify({"error": "已有任务正在运行"}), 400

    data = request.get_json(force=True)
    videos = data.get("videos", [])
    summary_type = data.get("summary_type", "game_guide")
    combine = bool(data.get("combine", False))
    if summary_type not in ("game_guide", "ai_interview", "wechat_material"):
        return jsonify({"error": "summary_type 只能是 game_guide / ai_interview / wechat_material（攻略成品文章走自媒体工作台生成）"}), 400
    if not videos:
        return jsonify({"error": "请选择要整理的视频"}), 400

    _task_status = {
        "running": True,
        "progress": "准备生成内容...",
        "done": False,
        "error": "",
        "success": 0,
        "summary_ids": [],
    }

    def run():
        global _task_status
        try:
            _task_status["progress"] = "同步 Excel 到内容库..."
            _sync_content_db()
            config = _ai_config()
            success_count = 0
            summary_ids = []
            selected_videos = []

            for i, key in enumerate(videos):
                try:
                    sheet, row_text = key.split(":", 1)
                    row = int(row_text)
                    _task_status["progress"] = f"读取 [{i+1}/{len(videos)}] {sheet} 第{row}行..."
                    with _db_lock:
                        video = content_store.get_video_by_source(DB_PATH, sheet, row)
                    if not video:
                        _task_status["error"] += f"{key}: 内容库中找不到视频\n"
                        continue
                    if not (video.get("transcript") or "").strip():
                        _task_status["error"] += f"{key}: 没有 ASR 文本，跳过\n"
                        continue
                    selected_videos.append(video)
                except Exception as e:
                    _task_status["error"] += f"{key}: {e}\n"

            if combine:
                if not selected_videos:
                    _task_status["progress"] = "没有可生成的 ASR 文本"
                    _task_status["done"] = True
                    return

                _task_status["progress"] = f"合并整理 {len(selected_videos)} 条视频..."
                result, model, status = content_store.generate_collection_summary(
                    selected_videos, summary_type, config
                )
                if summary_type in GUIDE_SUMMARY_TYPES:
                    # 攻略类：抽关键帧做配图，嵌入整理稿尾部
                    _task_status["progress"] = "抽取视频关键帧配图..."
                    result["content"] = (result.get("content") or "") + _video_keyframe_markdown(selected_videos)
                source_ids = [int(v["id"]) for v in selected_videos]
                with _db_lock:
                    summary_id = content_store.save_summary(
                        DB_PATH,
                        video_id=source_ids[0],
                        summary_type=summary_type,
                        result=result,
                        model=model,
                        status=status,
                        source_video_ids=source_ids,
                    )
                summary_ids.append(summary_id)
                success_count = 1
            else:
                for i, video in enumerate(selected_videos):
                    _task_status["progress"] = f"整理 [{i+1}/{len(selected_videos)}] 第{video['source_row']}行..."

                    result, model, status = content_store.generate_summary(
                        video, summary_type, config
                    )
                    if summary_type in GUIDE_SUMMARY_TYPES:
                        # 攻略类：抽关键帧做配图，嵌入整理稿尾部
                        _task_status["progress"] = f"抽取关键帧配图 [{i+1}/{len(selected_videos)}] 第{video['source_row']}行..."
                        result["content"] = (result.get("content") or "") + _video_keyframe_markdown([video])
                    with _db_lock:
                        summary_id = content_store.save_summary(
                            DB_PATH,
                            video_id=video["id"],
                            summary_type=summary_type,
                            result=result,
                            model=model,
                            status=status,
                            source_video_ids=[int(video["id"])],
                        )
                    summary_ids.append(summary_id)
                    success_count += 1

            _task_status["success"] = success_count
            _task_status["summary_ids"] = summary_ids
            label = "合并整理稿" if combine else "整理稿"
            _task_status["progress"] = f"✅ 已生成 {success_count} 条{label}"
            _task_status["done"] = True
        except Exception:
            _task_status["error"] = traceback.format_exc()
            _task_status["done"] = True
        finally:
            _task_status["running"] = False

    threading.Thread(target=run, daemon=True).start()
    return jsonify({"status": "started", "total": len(videos), "summary_type": summary_type, "combine": combine})


# ── 攻略整理稿配图：抽取视频关键帧并生成 Markdown 图片段 ──
# 成品文章在自媒体工作台生成；这里只给资料整理稿（game_guide / wechat_material 公众号素材档案）配图，供工作台取材时引用
GUIDE_SUMMARY_TYPES = ("game_guide", "wechat_material")

_extractor = None
_extractor_lock = threading.Lock()


def _load_cached_nwm_url(aweme_id: str) -> str:
    """读采集时缓存的原始解析结果（output/raw/douyin/<id>.json），免网络请求。"""
    import json as _json
    path = material_store.RAW_DIR / f"{material_store.safe_id(aweme_id)}.json"
    try:
        if path.exists():
            raw = _json.loads(path.read_text(encoding="utf-8"))
            return (raw.get("nwm_url") or raw.get("download_url") or "").strip()
    except Exception:
        pass
    return ""


def _resolve_download_url(video: dict, retries: int = 4, backoff: float = 15.0) -> str:
    """解析视频真实无水印直链：本地 raw 缓存优先，其次 douyin_video_parser 在线解析。

    Excel 里的 video_url 是抖音网页地址，curl 下载只会拿到 HTML；必须解析出
    nwm 直链 ffmpeg 才能抽帧。注意：
    1) parser 构造时按相对路径读 douyin_cookie.txt（依赖 cwd），需显式 set_cookie 注入
    2) 抖音接口有间歇性反爬限流，失败需带退避重试
    解析失败返回空串，由调用方兜底原地址。
    """
    aweme_id = (video.get("aweme_id") or "").strip()
    if not aweme_id:
        return ""
    cached = _load_cached_nwm_url(aweme_id)
    if cached:
        return cached
    global _extractor
    with _extractor_lock:
        if _extractor is None:
            try:
                from src.video_extractor import VideoExtractor
                _extractor = VideoExtractor()
                cookie = get_cookie_path().read_text(encoding="utf-8").strip()
                if cookie:
                    _extractor.parser.set_cookie(cookie)
            except Exception:
                traceback.print_exc()
                return ""
        for attempt in range(retries):
            try:
                # parse_video 的 get_video_id 只认 URL 形态，裸 aweme_id 会直接返回 None
                url = (video.get("source_url") or "").strip() or (video.get("video_url") or "").strip()
                if "douyin.com" not in url:
                    url = f"https://www.douyin.com/video/{aweme_id}"
                info = _extractor.parser.parse_video(url)
                nwm = (info.get("nwm_url") or "").strip() if info else ""
                if nwm:
                    # 解析成功即落 raw 缓存：抖音接口有短窗限流，缓存后不再走网络
                    try:
                        material_store.save_raw(aweme_id, {"nwm_url": nwm})
                    except Exception:
                        pass
                    return nwm
            except Exception:
                pass
            if attempt < retries - 1:
                time.sleep(backoff * (attempt + 1))
    return ""


def _video_keyframe_markdown(videos: list, max_frames: int = 5, per_video: int = 2) -> str:
    """对攻略类整理稿的视频抽关键帧（ffmpeg，已有帧则直接复用），返回 Markdown 配图段。

    单视频模式取最多 max_frames=5 张；合并模式每视频最多 per_video=2 张、总数截到 max_frames。
    抽帧失败（无视频地址/下载失败）时返回空串，不影响整理稿本身。
    """
    single = len(videos) <= 1
    collected: list[str] = []
    for video in videos:
        if len(collected) >= max_frames:
            break
        aweme_id = (video.get("aweme_id") or "").strip() or material_store.extract_aweme_id(
            video.get("source_url") or video.get("video_url") or ""
        )
        if not aweme_id:
            continue
        # 已有帧直接复用，跳过解析下载
        if not material_store.local_keyframe_paths(aweme_id):
            video_url = (video.get("video_url") or "").strip()
            nwm_url = _resolve_download_url(video)
            if nwm_url:
                material_store.extract_keyframes(aweme_id, nwm_url, max_frames=max_frames)
            elif video_url:
                material_store.extract_keyframes(aweme_id, video_url, max_frames=max_frames)
        frame_limit = per_video if not single else max_frames
        for rel in material_store.local_keyframe_paths(aweme_id):
            if rel.endswith("sheet.jpg"):
                continue
            if len(collected) >= max_frames or frame_limit <= 0:
                break
            collected.append(rel)
            frame_limit -= 1
        if single:
            break
    if not collected:
        return ""
    lines = ["", "## 配图素材（视频关键帧）", ""]
    for idx, rel in enumerate(collected, 1):
        lines.append(f"![关键帧{idx}](/media/{rel})")
    lines.append("")
    return "\n".join(lines)


# ── AI 配置管理 ──
ENV_PATH = ROOT / "config" / "config.env.local"


def _mask_key(key: str) -> str:
    if not key:
        return ""
    if len(key) <= 10:
        return "***"
    return f"{key[:5]}...{key[-4:]}"


def _update_env_file(path: Path, updates: dict):
    """按行更新 KEY=VALUE，保留原有注释与其他键；新键追加到末尾。"""
    lines = []
    if path.exists():
        lines = path.read_text(encoding="utf-8").splitlines()
    remaining = dict(updates)
    out = []
    for line in lines:
        stripped = line.strip()
        if stripped and not stripped.startswith("#") and "=" in stripped:
            key = stripped.split("=", 1)[0].strip()
            if key in remaining:
                out.append(f"{key}={remaining.pop(key)}")
                continue
        out.append(line)
    for key, val in remaining.items():
        out.append(f"{key}={val}")
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text("\n".join(out) + "\n", encoding="utf-8")


@app.route("/api/ai/config", methods=["GET"])
def api_ai_config_get():
    config = collector.load_env_file(str(ENV_PATH))
    method = collector.normalize_ai_method(collector.config_get(config, "AI_METHOD", "skip"))
    if method == "deepseek":
        key = collector.config_get(config, "DEEPSEEK_API_KEY") or collector.config_get(config, "AI_API_KEY")
        base = collector.config_get(config, "DEEPSEEK_API_BASE", "")
        model = collector.config_get(config, "DEEPSEEK_MODEL", "")
    else:
        key = collector.config_get(config, "OPENAI_API_KEY") or collector.config_get(config, "AI_API_KEY")
        base = collector.config_get(config, "AI_API_BASE", "")
        model = collector.config_get(config, "AI_MODEL", "")
    return jsonify({
        "method": method,
        "has_key": bool(key),
        "api_key_masked": _mask_key(key),
        "api_base": base,
        "model": model,
    })


@app.route("/api/ai/config", methods=["POST"])
def api_ai_config_save():
    data = request.get_json(force=True)
    method = collector.normalize_ai_method(data.get("method", "skip"))
    if method not in ("skip", "deepseek", "openai"):
        return jsonify({"error": "AI 服务类型无效"}), 400

    config = collector.load_env_file(str(ENV_PATH))
    updates = {"AI_METHOD": method}
    if method != "skip":
        api_key = (data.get("api_key") or "").strip()
        if not api_key:
            # 表单留空：沿用已保存的 Key
            api_key = config.get("DEEPSEEK_API_KEY" if method == "deepseek" else "OPENAI_API_KEY", "")
        if not api_key:
            return jsonify({"error": "请填写 API Key"}), 400
        base = (data.get("api_base") or "").strip()
        model = (data.get("model") or "").strip()
        if method == "deepseek":
            updates["DEEPSEEK_API_KEY"] = api_key
            updates["DEEPSEEK_API_BASE"] = base
            updates["DEEPSEEK_MODEL"] = model
        else:
            updates["OPENAI_API_KEY"] = api_key
            updates["AI_API_BASE"] = base
            updates["AI_MODEL"] = model
    _update_env_file(ENV_PATH, updates)
    return jsonify({"success": True})


def _ai_config_from_payload(data: dict) -> dict:
    """从请求体构建测试用配置；空字段回退到已保存配置或默认值。"""
    saved = _ai_config()
    method = collector.normalize_ai_method((data.get("method") or "").strip() or saved["method"])
    api_key = (data.get("api_key") or "").strip() or saved.get("api_key") or ""
    if method == "deepseek":
        default_base, default_model = "https://api.deepseek.com", "deepseek-chat"
    else:
        default_base, default_model = "https://api.openai.com/v1", "gpt-4o-mini"
    same_method = saved.get("method") == method
    api_base = (data.get("api_base") or "").strip() or (saved.get("api_base") if same_method else "") or default_base
    model = (data.get("model") or "").strip() or (saved.get("model") if same_method else "") or default_model
    return {"method": method, "api_key": api_key, "api_base": api_base, "model": model}


@app.route("/api/ai/test", methods=["POST"])
def api_ai_test():
    data = request.get_json(force=True) or {}
    cfg = _ai_config_from_payload(data)
    if cfg["method"] == "skip":
        return jsonify({"success": False, "error": "当前选择「不使用 AI」，请先选择 DeepSeek 或 OpenAI 接口"})
    if not cfg["api_key"]:
        return jsonify({"success": False, "error": "请填写 API Key"})
    try:
        import openai
    except ImportError:
        return jsonify({"success": False, "error": "服务端缺少 openai 库，请执行: pip3 install openai"})
    try:
        client = openai.OpenAI(api_key=cfg["api_key"], base_url=cfg["api_base"], timeout=30, max_retries=0)
        resp = client.chat.completions.create(
            model=cfg["model"],
            messages=[{"role": "user", "content": "连接测试，请只回复两个字母：OK"}],
            max_tokens=8,
        )
        reply = (resp.choices[0].message.content or "").strip()
        return jsonify({"success": True, "reply": reply, "model": cfg["model"]})
    except Exception as e:
        return jsonify({"success": False, "error": str(e)})


# ── Dify 知识库配置与发布 ──
def _dify_config() -> dict:
    config = collector.load_env_file(str(ENV_PATH))
    return {
        "api_base": collector.config_get(config, "DIFY_API_BASE", "http://127.0.0.1"),
        "api_key": collector.config_get(config, "DIFY_KB_API_KEY"),
        "dataset_id": collector.config_get(config, "DIFY_DATASET_ID"),
    }


@app.route("/api/dify/config", methods=["GET"])
def api_dify_config_get():
    cfg = _dify_config()
    return jsonify({
        "api_base": cfg["api_base"],
        "has_key": bool(cfg["api_key"]),
        "api_key_masked": _mask_key(cfg["api_key"]),
        "dataset_id": cfg["dataset_id"],
    })


@app.route("/api/dify/config", methods=["POST"])
def api_dify_config_save():
    data = request.get_json(force=True) or {}
    saved = _dify_config()
    updates = {}
    api_base = (data.get("api_base") or "").strip()
    if api_base:
        updates["DIFY_API_BASE"] = api_base
    api_key = (data.get("api_key") or "").strip()
    if api_key:
        updates["DIFY_KB_API_KEY"] = api_key  # 留空则沿用已保存 Key
    dataset_id = (data.get("dataset_id") or "").strip()
    if dataset_id:
        updates["DIFY_DATASET_ID"] = dataset_id
    if not updates:
        return jsonify({"error": "没有需要保存的配置"}), 400
    _update_env_file(ENV_PATH, updates)
    return jsonify({"success": True})


@app.route("/api/dify/datasets")
def api_dify_datasets():
    cfg = _dify_config()
    if not cfg["api_key"]:
        return jsonify({"error": "请先填写 Dify Knowledge API Key"}), 400
    try:
        client = DifyKBClient(cfg["api_base"], cfg["api_key"], timeout=20)
        datasets = client.list_datasets()
        return jsonify({"datasets": [
            {"id": d.get("id"), "name": d.get("name"),
             "doc_count": d.get("document_count", 0)} for d in datasets
        ]})
    except DifyKBError as e:
        return jsonify({"error": str(e)}), 502
    except Exception as e:
        return jsonify({"error": str(e)}), 500


@app.route("/api/dify/datasets/create", methods=["POST"])
def api_dify_datasets_create():
    data = request.get_json(force=True) or {}
    name = (data.get("name") or "").strip()
    if not name:
        return jsonify({"error": "请填写知识库名称"}), 400
    cfg = _dify_config()
    if not cfg["api_key"]:
        return jsonify({"error": "请先填写 Dify Knowledge API Key"}), 400
    try:
        client = DifyKBClient(cfg["api_base"], cfg["api_key"], timeout=30)
        ds = client.create_dataset(name)
        _update_env_file(ENV_PATH, {"DIFY_DATASET_ID": ds.get("id", "")})
        return jsonify({"success": True, "id": ds.get("id"), "name": ds.get("name")})
    except DifyKBError as e:
        return jsonify({"error": str(e)}), 502
    except Exception as e:
        return jsonify({"error": str(e)}), 500


@app.route("/api/dify/publish", methods=["POST"])
def api_dify_publish():
    if _dify_status.get("running"):
        return jsonify({"error": "发布任务进行中，请稍候"}), 409
    cfg = _dify_config()
    if not cfg["api_key"]:
        return jsonify({"error": "请先在 AI 设置中配置 Dify Knowledge API Key"}), 400
    data = request.get_json(force=True, silent=True) or {}
    keys = data.get("keys") or []
    force = bool(data.get("force"))

    def run():
        _dify_status.update(running=True, done=False, error="", progress="准备发布...", result=None)
        try:
            client = DifyKBClient(cfg["api_base"], cfg["api_key"], timeout=90)
            with _db_lock:
                candidates = content_store.get_dify_candidates(DB_PATH, keys or None)
            if not force and keys:
                candidates = [v for v in candidates if not v.get("dify_document_id")] or candidates
            total = len(candidates)
            ok = fail = 0
            errors = []
            for i, video in enumerate(candidates, 1):
                _dify_status["progress"] = f"发布 {i}/{total}：{(video.get('title') or '')[:30]}"
                try:
                    name, text = content_store.build_dify_document(video)
                    doc_id = video.get("dify_document_id")
                    if doc_id:
                        client.update_document_by_text(cfg["dataset_id"], doc_id, name, text)
                    else:
                        resp = client.create_document_by_text(cfg["dataset_id"], name, text)
                        doc_id = (resp.get("document") or {}).get("id") or resp.get("id")
                    if not doc_id:
                        raise DifyKBError("未返回 document id")
                    with _db_lock:
                        content_store.set_dify_doc(
                            DB_PATH, video["id"], doc_id, datetime.now().strftime("%Y-%m-%d %H:%M")
                        )
                    ok += 1
                except (DifyKBError, Exception) as e:
                    fail += 1
                    errors.append(f'{(video.get("title") or "未命名")[:24]}: {e}')
            _dify_status["result"] = {"ok": ok, "fail": fail, "total": total, "errors": errors[:10]}
            _dify_status["progress"] = f"完成：成功 {ok} / 失败 {fail}"
        except Exception as e:
            _dify_status["error"] = f"{e}\n{traceback.format_exc()[-500:]}"
        finally:
            _dify_status["running"] = False
            _dify_status["done"] = True

    threading.Thread(target=run, daemon=True).start()
    return jsonify({"success": True})


@app.route("/api/dify/publish/status")
def api_dify_publish_status():
    return jsonify(_dify_status)


# ════════════════════════════════════════════════════════════════
# 订阅（作者主页增量同步）
# ════════════════════════════════════════════════════════════════

_sub_status = {"running": False, "progress": "", "done": True, "error": "", "result": None}


def _read_cookie_from_file() -> str:
    """读本地 cookie 文件（采集中心每次采集都会写入）。"""
    try:
        p = get_cookie_path()
        if p.exists():
            return p.read_text(encoding="utf-8").strip()
    except Exception:
        pass
    return ""


@app.route("/api/subscriptions")
def api_subscriptions_list():
    if not DB_PATH.exists():
        return jsonify({"items": []})
    try:
        with _db_lock:
            items = content_store.list_subscriptions(DB_PATH)
    except Exception as e:
        return jsonify({"items": [], "error": str(e)}), 500
    return jsonify({"items": items})


@app.route("/api/subscriptions", methods=["POST"])
def api_subscriptions_add():
    data = request.get_json(force=True)
    url = (data.get("url") or "").strip()
    author = (data.get("author") or "").strip()
    category = (data.get("category") or "").strip()
    game = (data.get("game") or "").strip()
    platform = (data.get("platform") or "douyin").strip()
    if not url:
        return jsonify({"error": "请输入链接"}), 400

    if platform == "wechat":
        # 微信：url 是 WeWe RSS 的 feed URL
        if not url.startswith("http"):
            return jsonify({"error": "RSS 链接需以 http 开头"}), 400
        sec_uid = f"wechat:{hashlib.md5(url.encode()).hexdigest()[:12]}"
        try:
            with _db_lock:
                sub = content_store.add_subscription(
                    DB_PATH, sec_uid, url, author, category, game, "wechat")
        except Exception as e:
            return jsonify({"error": str(e)}), 500
        return jsonify({"success": True, "subscription": sub})

    # 抖音：url 是用户主页链接
    from src.fetch_user_videos import extract_sec_uid_from_url, resolve_short_url
    if "v.douyin.com" in url or "/share/" in url:
        url = resolve_short_url(url)
    sec_uid = extract_sec_uid_from_url(url)
    if not sec_uid:
        return jsonify({"error": "无法从链接提取 sec_uid，请使用 www.douyin.com/user/... 格式"}), 400
    user_url = f"https://www.douyin.com/user/{sec_uid}"
    try:
        with _db_lock:
            sub = content_store.add_subscription(
                DB_PATH, sec_uid, user_url, author, category, game, "douyin")
    except Exception as e:
        return jsonify({"error": str(e)}), 500
    return jsonify({"success": True, "subscription": sub})


@app.route("/api/authors")
def api_authors():
    """视频库作者统计（含是否已订阅）。"""
    if not DB_PATH.exists():
        return jsonify({"authors": []})
    try:
        with _db_lock:
            conn = content_store.connect(DB_PATH)
            try:
                rows = conn.execute(
                    "SELECT COALESCE(NULLIF(TRIM(author),''),'未知作者') AS a, COUNT(*) AS n "
                    "FROM videos GROUP BY a ORDER BY n DESC"
                ).fetchall()
                subs = {s["author"] for s in content_store.list_subscriptions(DB_PATH)}
            finally:
                conn.close()
        authors = [
            {"author": r["a"], "count": r["n"], "subscribed": r["a"] in subs}
            for r in rows
        ]
        return jsonify({"authors": authors})
    except Exception as e:
        return jsonify({"authors": [], "error": str(e)}), 500


@app.route("/api/subscriptions/import", methods=["POST"])
def api_subscriptions_import():
    """从视频库作者导入订阅：拿该作者一条视频链接反查 sec_uid。"""
    data = request.get_json(force=True)
    author = (data.get("author") or "").strip()
    if not author:
        return jsonify({"error": "缺少作者名"}), 400
    cookie = _read_cookie_from_file()
    if not cookie:
        return jsonify({"error": "未找到 Cookie，请先在采集中心完成一次采集"}), 400

    # 找该作者一条已完成视频的链接
    try:
        with _db_lock:
            conn = content_store.connect(DB_PATH)
            try:
                row = conn.execute(
                    "SELECT source_url FROM videos "
                    "WHERE author = ? AND COALESCE(source_url,'') != '' "
                    "ORDER BY id DESC LIMIT 1", (author,),
                ).fetchone()
            finally:
                conn.close()
    except Exception as e:
        return jsonify({"error": str(e)}), 500
    if not row or not row["source_url"]:
        return jsonify({"error": f"视频库中没有「{author}」的视频链接"}), 404

    try:
        parser = _setup_parser_and_cookie(cookie)
        home = parser.get_user_home_from_video_url(row["source_url"])
        if not home:
            return jsonify({"error": "无法从视频链接解析作者主页，链接可能已失效"}), 502
        from src.fetch_user_videos import extract_sec_uid_from_url
        sec_uid = extract_sec_uid_from_url(home)
        if not sec_uid:
            return jsonify({"error": "解析主页失败：无法提取 sec_uid"}), 502
        with _db_lock:
            sub = content_store.add_subscription(DB_PATH, sec_uid, home, author)
        return jsonify({"success": True, "subscription": sub})
    except Exception as e:
        return jsonify({"error": str(e)}), 500


@app.route("/api/subscriptions/<int:sub_id>/author", methods=["POST"])
def api_subscriptions_author(sub_id):
    """修改订阅作者显示名。"""
    data = request.get_json(force=True)
    author = (data.get("author") or "").strip()
    if not author:
        return jsonify({"error": "作者名不能为空"}), 400
    try:
        with _db_lock:
            ok = content_store.update_subscription_author(DB_PATH, sub_id, author)
    except Exception as e:
        return jsonify({"error": str(e)}), 500
    if not ok:
        return jsonify({"error": "订阅不存在"}), 404
    return jsonify({"success": True})


@app.route("/api/subscriptions/<int:sub_id>/category", methods=["POST"])
def api_subscriptions_category(sub_id):
    """修改订阅作者的分类（同步的新视频会自动打上）。"""
    data = request.get_json(force=True)
    category = (data.get("category") or "").strip()
    game = (data.get("game") or "").strip()
    if category not in content_store.CATEGORIES:
        return jsonify({"error": "无效分类"}), 400
    try:
        with _db_lock:
            ok = content_store.update_subscription_category(DB_PATH, sub_id, category, game)
    except Exception as e:
        return jsonify({"error": str(e)}), 500
    if not ok:
        return jsonify({"error": "订阅不存在"}), 404
    return jsonify({"success": True})


@app.route("/api/subscriptions/<int:sub_id>", methods=["DELETE"])
def api_subscriptions_delete(sub_id):
    try:
        with _db_lock:
            ok = content_store.delete_subscription(DB_PATH, sub_id)
    except Exception as e:
        return jsonify({"error": str(e)}), 500
    if not ok:
        return jsonify({"error": "订阅不存在"}), 404
    return jsonify({"success": True})


@app.route("/api/subscriptions/sync", methods=["POST"])
def api_subscriptions_sync():
    """同步订阅（全部/单个/多个勾选）。抖音：拉主页新视频→ASR。微信：拉 RSS 新文章。"""
    if _sub_status.get("running"):
        return jsonify({"error": "订阅同步进行中，请稍候"}), 409
    data = request.get_json(force=True, silent=True) or {}
    sub_id = data.get("id")  # None = 全部
    sub_ids = data.get("ids") or []  # 勾选批量：[1, 3, 7]；空 = 不启用批量过滤
    cookie = (data.get("cookie") or "").strip() or _read_cookie_from_file()

    def run():
        import time
        _sub_status.update(running=True, done=False, error="", progress="准备同步...", result=None)
        try:
            with _db_lock:
                subs = content_store.list_subscriptions(DB_PATH)
            if sub_id:
                subs = [s for s in subs if s["id"] == sub_id]
            elif sub_ids:
                wanted = {int(i) for i in sub_ids if str(i).isdigit() or isinstance(i, int)}
                subs = [s for s in subs if s["id"] in wanted]
            if not subs:
                _sub_status["error"] = "没有可同步的订阅"
                return

            # 分流：抖音需要 cookie + parser；微信不需要
            douyin_subs = [s for s in subs if s.get("platform", "douyin") == "douyin"]
            wechat_subs = [s for s in subs if s.get("platform") == "wechat"]

            if douyin_subs and not cookie:
                _sub_status["error"] = "抖音订阅需要 Cookie，请先在采集中心完成一次采集"
                return

            summary = []

            # ── 微信订阅同步 ──
            if wechat_subs:
                from src.wechat_fetcher import sync_wechat_feed
                for si, sub in enumerate(wechat_subs, 1):
                    name = sub["author"] or sub["sec_uid"][:16]
                    _sub_status["progress"] = f"[微信 {si}/{len(wechat_subs)}] {name}: 拉取 RSS..."
                    try:
                        existing = content_store.get_wechat_article_links(DB_PATH)
                        result = sync_wechat_feed(sub["user_url"], existing)
                        if result["error"]:
                            with _db_lock:
                                content_store.update_subscription_sync(
                                    DB_PATH, sub["id"], 0, result["error"])
                            summary.append(f"{name}: {result['error']}")
                            continue
                        new_articles = result["articles"]
                        if not new_articles:
                            with _db_lock:
                                content_store.update_subscription_sync(
                                    DB_PATH, sub["id"], 0, "无新文章")
                            summary.append(f"{name}: 无新文章")
                            continue
                        with _db_lock:
                            inserted = content_store.insert_wechat_articles(
                                DB_PATH, new_articles,
                                category=sub.get("category") or "",
                                game=sub.get("game") or "",
                            )
                        # 更新索引
                        collector.update_video_index(str(EXCEL_PATH))
                        _sync_content_db()
                        cat_note = f"，已按订阅标记分类「{sub['category']}」{len(inserted)} 条" if sub.get("category") and inserted else ""
                        with _db_lock:
                            content_store.update_subscription_sync(
                                DB_PATH, sub["id"], len(inserted),
                                f"新增 {len(inserted)} 篇文章{cat_note}",
                                author=sub.get("author") or "")
                        summary.append(f"{name}: 新增 {len(inserted)} 篇文章{cat_note}")
                    except Exception as e:
                        with _db_lock:
                            content_store.update_subscription_sync(
                                DB_PATH, sub["id"], 0, f"同步失败: {e}")
                        summary.append(f"{name}: 失败 ({e})")

            # ── 抖音订阅同步 ──
            if douyin_subs:
                parser = _setup_parser_and_cookie(cookie)
                ai_config = {"method": "skip", "api_key": "", "api_base": "", "model": ""}
                for si, sub in enumerate(douyin_subs, 1):
                    name = sub["author"] or sub["sec_uid"][:12]
                    _sub_status["progress"] = f"[抖音 {si}/{len(douyin_subs)}] {name}: 拉取主页..."
                    try:
                        from src.fetch_user_videos import fetch_user_videos
                        result = fetch_user_videos(
                            url=sub["user_url"], cookie=cookie,
                            max_pages=1, max_videos=20, mode="user_url",
                            exclude_excel=str(EXCEL_PATH),
                        )
                        videos = result.get("videos") or []
                        new_author = ""
                        if not videos:
                            with _db_lock:
                                content_store.update_subscription_sync(
                                    DB_PATH, sub["id"], 0, "无新视频")
                            summary.append(f"{name}: 无新视频")
                            continue

                        # 写入 Excel
                        import openpyxl
                        wb = openpyxl.load_workbook(str(EXCEL_PATH))
                        ws = wb["抖音视频数据"]
                        next_row = ws.max_row + 1
                        for r in range(2, ws.max_row + 1):
                            if not ws.cell(r, 1).value:
                                next_row = r
                                break
                        for i, v in enumerate(videos):
                            ws.cell(next_row + i, 1).value = v["url"]
                            ws.cell(next_row + i, 2).value = "未开始"
                            ws.cell(next_row + i, 3).value = v.get("aweme_id", "")
                        wb.save(str(EXCEL_PATH))
                        wb.close()

                        # 逐条 ASR
                        for i, v in enumerate(videos):
                            row = next_row + i
                            _sub_status["progress"] = (
                                f"[抖音 {si}/{len(douyin_subs)}] {name}: 处理 {i+1}/{len(videos)} "
                                f"{v.get('aweme_id','')}..."
                            )
                            wb = openpyxl.load_workbook(str(EXCEL_PATH))
                            ws = wb["抖音视频数据"]
                            ok = collector.process_row(ws, row, cookie, "base", ai_config, parser)
                            if not new_author:
                                new_author = str(ws.cell(row, 4).value or "").strip()
                            wb.save(str(EXCEL_PATH))
                            wb.close()
                            if i < len(videos) - 1:
                                time.sleep(5)

                        # 更新索引 + 内容库 + 订阅状态
                        collector.update_video_index(str(EXCEL_PATH))
                        _sync_content_db()
                        # 订阅设了分类 → 新视频自动打上（AI 分类只补 category 为空的，不会覆盖）
                        cat_note = ""
                        if sub.get("category"):
                            with _db_lock:
                                conn = content_store.connect(DB_PATH)
                                try:
                                    marked = 0
                                    for i in range(len(videos)):
                                        r = conn.execute(
                                            "SELECT id FROM videos WHERE source_sheet = ? AND source_row = ?",
                                            ("抖音视频数据", next_row + i),
                                        ).fetchone()
                                        if r:
                                            # 非游戏攻略分类不带 game
                                            g = sub.get("game") or ""
                                            if sub["category"] != "游戏攻略":
                                                g = ""
                                            conn.execute(
                                                "UPDATE videos SET category = ?, game = ? WHERE id = ?",
                                                (sub["category"], g, r["id"]),
                                            )
                                            marked += 1
                                    conn.commit()
                                finally:
                                    conn.close()
                            cat_note = f"，已按订阅标记分类「{sub['category']}」{marked} 条"
                        with _db_lock:
                            content_store.update_subscription_sync(
                                DB_PATH, sub["id"], len(videos),
                                f"新增 {len(videos)} 条并已转写{cat_note}", author=new_author)
                        summary.append(f"{name}: 新增 {len(videos)} 条{cat_note}")
                    except Exception as e:
                        with _db_lock:
                            content_store.update_subscription_sync(
                                DB_PATH, sub["id"], 0, f"同步失败: {e}")
                        summary.append(f"{name}: 失败 ({e})")

            _sub_status["result"] = summary
            _sub_status["progress"] = "✅ 订阅同步完成：" + "；".join(summary)
        except Exception as e:
            _sub_status["error"] = f"{e}\n{traceback.format_exc()[-500:]}"
        finally:
            _sub_status["running"] = False
            _sub_status["done"] = True

    threading.Thread(target=run, daemon=True).start()
    return jsonify({"success": True})


@app.route("/api/subscriptions/sync/status")
def api_subscriptions_sync_status():
    return jsonify(_sub_status)


# ════════════════════════════════════════════════════════════════
# 删除：视频 / 面试题
# ════════════════════════════════════════════════════════════════

@app.route("/api/videos/delete", methods=["POST"])
def api_videos_delete():
    """删除视频：清 Excel 行（保持行号稳定）→ 删 SQLite 记录 → 重建索引。"""
    data = request.get_json(force=True)
    items = data.get("items") or []
    if not items:
        return jsonify({"error": "请选择要删除的视频"}), 400
    try:
        import openpyxl

        deleted = 0
        with _db_lock:
            # 1. 删 SQLite
            for it in items:
                sheet, row = it.get("sheet"), int(it.get("row") or 0)
                if sheet and row:
                    content_store.delete_video(DB_PATH, sheet, row)
            # 2. 清 Excel 行（所有 sheet 中匹配行清空 1-15 列，行号保持不变）
            wb = openpyxl.load_workbook(str(EXCEL_PATH))
            target_sheets = {it.get("sheet") for it in items if it.get("sheet")}
            keyset = {(it.get("sheet"), int(it.get("row") or 0)) for it in items}
            for ws in wb.worksheets:
                if ws.title not in target_sheets:
                    continue
                for (s, r) in keyset:
                    if s == ws.title and 2 <= r <= ws.max_row:
                        for col in range(1, 16):
                            ws.cell(r, col).value = None
                        deleted += 1
            wb.save(str(EXCEL_PATH))
            wb.close()
            # 3. 重建索引（追加分）+ 显式移除已删条目（索引是合并模式，不会自动删）
            collector.update_video_index(str(EXCEL_PATH))
            if INDEX_PATH.exists():
                with open(INDEX_PATH, encoding="utf-8") as f:
                    idx = json.load(f)
                idx["videos"] = [
                    v for v in idx.get("videos", [])
                    if (v.get("sheet"), int(v.get("row") or 0)) not in keyset
                ]
                with open(INDEX_PATH, "w", encoding="utf-8") as f:
                    json.dump(idx, f, ensure_ascii=False, indent=2)
            # 4. 重同步内容库
            _sync_content_db()
        return jsonify({"success": True, "deleted": deleted})
    except Exception as e:
        return jsonify({"error": str(e)}), 500


@app.route("/api/content/questions/delete", methods=["POST"])
def api_content_question_delete():
    """删除面试题库中的单题（改写 structured_data.questions）。"""
    data = request.get_json(force=True)
    summary_id = int(data.get("summary_id") or 0)
    q_index = int(data.get("q_index") or -1)
    if not summary_id or q_index < 0:
        return jsonify({"error": "参数错误"}), 400
    try:
        with _db_lock:
            conn = content_store.connect(DB_PATH)
            try:
                row = conn.execute(
                    "SELECT structured_data FROM ai_summaries WHERE id = ? AND summary_type = 'ai_interview'",
                    (summary_id,),
                ).fetchone()
                if not row:
                    return jsonify({"error": "面试题整理稿不存在"}), 404
                structured = json.loads(row["structured_data"] or "{}")
                questions = structured.get("questions") or []
                if q_index >= len(questions):
                    return jsonify({"error": "题目索引越界"}), 400
                questions.pop(q_index)
                structured["questions"] = questions
                conn.execute(
                    "UPDATE ai_summaries SET structured_data = ? WHERE id = ?",
                    (json.dumps(structured, ensure_ascii=False), summary_id),
                )
                conn.commit()
                remaining = len(questions)
            finally:
                conn.close()
        # 全删完则连整理稿一起删
        if remaining == 0:
            content_store.delete_summary(DB_PATH, summary_id)
        return jsonify({"success": True, "remaining": remaining})
    except Exception as e:
        return jsonify({"error": str(e)}), 500


@app.route("/")
def index():
    return send_from_directory(str(ROOT / "web" / "templates"), "index.html")

if __name__ == "__main__":
    print(f"🎬 video2text Web 面板")
    print(f"   Excel: {EXCEL_PATH}")
    print(f"   启动: http://127.0.0.1:15801")
    app.run(host="127.0.0.1", port=15801, debug=False)
