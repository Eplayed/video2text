#!/usr/bin/env python3
"""
抖音视频采集器 — 主程序
用法：
  python main.py --excel /path/to/抖音视频信息.xlsx --cookie "sessionid=xxx"
  python main.py --excel /path/to/抖音视频信息.xlsx --cookie-file /path/to/cookie.txt
"""

import argparse
import os
import sys
import re
import time
import json
import tempfile
import subprocess
from datetime import datetime
from pathlib import Path

# 第三方库
try:
    import openpyxl
except ImportError:
    print("需要 openpyxl: pip install openpyxl")
    sys.exit(1)

# faster-whisper（需 KMP_DUPLICATE_LIB_OK=TRUE）
try:
    from faster_whisper import WhisperModel
    HAS_FASTER_WHISPER = True
except ImportError:
    HAS_FASTER_WHISPER = False
    import whisper

# ── 路径配置 ──
sys.path.insert(0, str(Path(__file__).resolve().parent))
from src.path_config import find_parser_dir, get_cookie_path

# ── Excel 列索引（1-based）───────────────────────────────────────
COL = {
    "链接": 1, "状态": 2, "视频ID": 3, "作者": 4, "发布时间": 5,
    "标题": 6, "视频文案ASR": 7, "标签": 8, "封面URL": 9,
    "视频链接": 10, "口播原文": 11, "AI优化文案": 12,
    "AI备选标题": 13, "关键词摘要": 14, "备注": 15,
}

STATUS = {"未开始": "未开始", "处理中": "处理中",
          "已完成": "已完成", "已写文稿": "已写文稿",
          "需核查": "需核查", "已废弃": "已废弃", "失败": "失败"}


def load_env_file(path: str) -> dict:
    """读取简单 KEY=VALUE 配置文件，不覆盖系统环境变量。"""
    env = {}
    if not path or not os.path.exists(path):
        return env
    with open(path, encoding="utf-8") as f:
        for line in f:
            line = line.strip()
            if not line or line.startswith("#") or "=" not in line:
                continue
            key, value = line.split("=", 1)
            env[key.strip()] = value.strip().strip('"').strip("'")
    return env


def config_get(config: dict, key: str, default: str = "") -> str:
    return os.environ.get(key) or config.get(key) or default


def normalize_ai_method(method: str) -> str:
    aliases = {
        "openai_api": "openai",
        "deepseek_api": "deepseek",
        "openai": "openai",
        "deepseek": "deepseek",
        "skip": "skip",
    }
    return aliases.get((method or "skip").strip(), "skip")


def get_douyin_parser(parser_dir: str):
    """从可配置目录加载 douyin_parse，避免把 /tmp 路径写死。"""
    parser_dir = os.path.abspath(parser_dir)
    if not os.path.isdir(parser_dir):
        raise RuntimeError(f"douyin_parse 目录不存在: {parser_dir}")
    if parser_dir not in sys.path:
        sys.path.insert(0, parser_dir)
    try:
        from douyin_video_parser import DouyinVideoParser
    except ImportError as e:
        raise RuntimeError(f"无法从 {parser_dir} 导入 douyin_video_parser: {e}") from e
    return DouyinVideoParser()


# ── 链接解析 ─────────────────────────────────────────────────────
def resolve_url(url: str) -> str:
    """跟随重定向，返回真实视频URL"""
    if not url or not isinstance(url, str):
        return ""
    url = url.strip()
    if not url:
        return ""
    try:
        import requests
        r = requests.head(url, timeout=10, allow_redirects=True,
                          headers={"User-Agent": "Mozilla/5.0"})
        return r.url
    except Exception:
        return url


def extract_aweme_id(url: str) -> str:
    """从URL中提取 aweme_id"""
    m = re.search(r"/video/(\d+)", url)
    if m:
        return m.group(1)
    return ""


# ── 视频元数据提取 ───────────────────────────────────────────────
def fetch_video_info(url_or_id: str, parser) -> dict:
    """调用 douyin_parse，返回结构化 dict"""
    try:
        raw = parser.parse_video(url_or_id)
        raw = raw or {}
        desc = raw.get("desc", "") or ""
        hashtags = []
        if "#" in desc:
            for part in desc.split("#")[1:]:
                tag = part.strip().split()[0]
                if tag:
                    hashtags.append(tag)

        ts = raw.get("create_time", 0)
        create_str = datetime.fromtimestamp(ts).strftime("%Y-%m-%d %H:%M:%S") if ts else ""

        video_obj = raw.get("video") or {}
        statistics = raw.get("statistics") or {}

        return {
            "aweme_id": raw.get("aweme_id", ""),
            "author": raw.get("author_nickname", ""),
            "author_sec_uid": raw.get("author_sec_uid", ""),
            "title": desc.split("#")[0].strip(),
            "desc": desc,
            "hashtags": " ".join(hashtags),
            "create_time": create_str,
            "duration": video_obj.get("duration"),
            "like_count": statistics.get("digg_count"),
            "comment_count": statistics.get("comment_count"),
            "share_count": statistics.get("share_count"),
            "collect_count": statistics.get("collect_count"),
            "cover_url": raw.get("cover_url", ""),
            "video_url": raw.get("nwm_url", ""),
            "real_url": f"https://www.douyin.com/video/{raw.get('aweme_id', '')}",
            "error": None,
        }
    except Exception as e:
        return {"error": str(e)}


# ── ASR 转写 ─────────────────────────────────────────────────────
# Whisper 模型缓存：同尺寸模型只加载一次（首次加载需数十秒，重复加载严重拖慢批量处理）
_ASR_MODELS = {}
_ASR_LOCK = __import__("threading").Lock()


def _get_asr_model(model_size: str):
    """获取（并缓存）指定尺寸的 Whisper 模型。"""
    with _ASR_LOCK:
        if model_size not in _ASR_MODELS:
            os.environ["KMP_DUPLICATE_LIB_OK"] = "TRUE"
            if HAS_FASTER_WHISPER:
                _ASR_MODELS[model_size] = WhisperModel(model_size, device="cpu", compute_type="int8")
            else:
                _ASR_MODELS[model_size] = whisper.load_model(model_size)
        return _ASR_MODELS[model_size]


def asr_transcribe(video_url: str, model_size: str = "base",
                    language: str = "zh") -> str:
    """下载视频 → 提取音频 → faster-whisper 转写"""
    tmp_dir = tempfile.mkdtemp(prefix="dy_")
    video_path = os.path.join(tmp_dir, "video.mp4")
    audio_path = os.path.join(tmp_dir, "audio.wav")

    # 下载视频
    try:
        subprocess.run(
            ["curl", "-L", "-o", video_path,
             "-H", "Referer: https://www.douyin.com/",
             video_url],
            capture_output=True, check=True, timeout=120
        )
    except Exception as e:
        return f"[下载失败] {e}"

    if not os.path.exists(video_path) or os.path.getsize(video_path) < 1024:
        return "[下载失败] 文件无效"

    # 提取音频
    try:
        subprocess.run(
            ["ffmpeg", "-i", video_path,
             "-vn", "-acodec", "pcm_s16le",
             "-ar", "16000", "-ac", "1",
             "-y", audio_path],
            capture_output=True, check=True, timeout=60
        )
    except Exception as e:
        return f"[音频提取失败] {e}"

    if not os.path.exists(audio_path):
        return "[音频提取失败] 文件不存在"

    # 转写
    try:
        if HAS_FASTER_WHISPER:
            model = _get_asr_model(model_size)
            segments, info = model.transcribe(
                audio_path, language=language,
                vad_filter=True, beam_size=1
            )
            text = "".join(s.text for s in segments)
        else:
            model = _get_asr_model(model_size)
            result = model.transcribe(audio_path, language=language)
            text = result["text"].strip()
    except Exception as e:
        return f"[转写失败] {e}"
    finally:
        # 清理临时文件
        for f in [video_path, audio_path]:
            if os.path.exists(f):
                os.remove(f)
        if os.path.exists(tmp_dir):
            os.rmdir(tmp_dir)

    return text


# ── AI 优化文案 ──────────────────────────────────────────────────
def ai_optimize(title: str, desc: str, author: str,
                transcript: str, config: dict) -> dict:
    """调用 LLM API 优化文案"""
    if config.get("method") == "skip" or not config.get("api_key"):
        return {}

    prompt = f"""你是短视频内容运营专家。请基于以下抖音视频信息，整理并优化文案。

标题：{title}
描述：{desc}
作者：{author}
口播识别文本：{transcript}

输出JSON（不要输出多余文字）：
{{
  "clean_transcript": "清洗后口播文案",
  "optimized_copy": "优化完整文案",
  "short_copy": "精简版100字内",
  "title_options": ["标题1","标题2","标题3","标题4","标题5"],
  "summary": "摘要50字内",
  "keywords": ["关键词1","关键词2"],
  "risk_words": ["敏感词"],
  "recommended_tags": ["标签1","标签2"]
}}"""

    try:
        import openai
        client = openai.OpenAI(
            api_key=config["api_key"],
            base_url=config.get("api_base", "https://api.openai.com/v1")
        )
        resp = client.chat.completions.create(
            model=config.get("model", "gpt-4o-mini"),
            messages=[{"role": "user", "content": prompt}],
            temperature=0.3,
            response_format={"type": "json_object"},
        )
        return json.loads(resp.choices[0].message.content)
    except Exception as e:
        return {"error": str(e)}


# ── 写 Excel ─────────────────────────────────────────────────────
def set_cell(ws, row: int, col_name: str, value):
    ws.cell(row, COL[col_name], value)


def process_row(ws, row_idx: int, cookie: str,
                asr_model: str, ai_config: dict,
                parser) -> bool:
    """处理单行，返回是否成功"""
    url = (ws.cell(row_idx, COL["链接"]).value or "").strip()
    if not url:
        return False

    # 标记处理中
    ws.cell(row_idx, COL["状态"], STATUS["处理中"])
    print(f"[Row {row_idx}] 处理中: {url[:60]}")

    # 1. 解析真实URL
    real_url = resolve_url(url)
    aweme_id = extract_aweme_id(real_url)
    if not aweme_id:
        ws.cell(row_idx, COL["状态"], STATUS["失败"])
        ws.cell(row_idx, COL["备注"], f"无法提取视频ID: {real_url}")
        return False

    # 2. 获取元数据（传完整URL而不是纯ID）
    info = fetch_video_info(real_url, parser)
    if info.get("error"):
        ws.cell(row_idx, COL["状态"], STATUS["失败"])
        ws.cell(row_idx, COL["备注"], info["error"])
        return False

    # 写入元数据列
    set_cell(ws, row_idx, "视频ID", info["aweme_id"])
    set_cell(ws, row_idx, "作者", info["author"])
    set_cell(ws, row_idx, "发布时间", info["create_time"])
    set_cell(ws, row_idx, "标题", info["title"])
    set_cell(ws, row_idx, "标签", info["hashtags"])
    set_cell(ws, row_idx, "封面URL", info["cover_url"])
    set_cell(ws, row_idx, "视频链接", info["real_url"])

    # 3. ASR 转写（如果还没有文案）
    existing_asr = (ws.cell(row_idx, COL["口播原文"]).value or "").strip()
    if not existing_asr:
        print(f"[Row {row_idx}] 开始ASR转写...")
        transcript = asr_transcribe(info["video_url"], model_size=asr_model)
        if transcript and not transcript.startswith("[") and not transcript.startswith("下载失败"):
            set_cell(ws, row_idx, "视频文案ASR", transcript)
            set_cell(ws, row_idx, "口播原文", transcript)
        else:
            ws.cell(row_idx, COL["备注"], transcript)
            print(f"[Row {row_idx}] ASR失败: {transcript}")

    # 4. AI 优化（如果配置了 API）
    if ai_config.get("api_key") and not ai_config.get("method") == "skip":
        existing = (ws.cell(row_idx, COL["AI优化文案"]).value or "").strip()
        if not existing:
            transcript = ws.cell(row_idx, COL["口播原文"]).value or ""
            result = ai_optimize(info["title"], info["desc"],
                                  info["author"], transcript, ai_config)
            if "error" not in result:
                set_cell(ws, row_idx, "AI优化文案", result.get("optimized_copy", ""))
                set_cell(ws, row_idx, "AI备选标题", " ".join(result.get("title_options", [])))
                set_cell(ws, row_idx, "关键词摘要", " ".join(result.get("keywords", [])))
            else:
                ws.cell(row_idx, COL["备注"], f"AI错误: {result.get('error')}")

    # 标记完成
    ws.cell(row_idx, COL["状态"], STATUS["已完成"])
    print(f"[Row {row_idx}] ✅ 完成")
    return True


# ── 主流程 ───────────────────────────────────────────────────────
def main():
    parser = argparse.ArgumentParser(description="抖音视频采集器")
    parser.add_argument("--sheet", default="抖音视频数据", help="Excel sheet名称")
    parser.add_argument("--excel", default="/Users/zhangyajun/Documents/project/video2text/output/抖音视频信息.xlsx")
    parser.add_argument("--config", default=str(Path(__file__).parent / "config/config.env.local"),
                        help="本地配置文件，默认 config/config.env.local")
    parser.add_argument("--parser-dir", default=find_parser_dir(),
                        help="douyin_parse 项目目录，自动查找 vendor/douyin_parse 或 DOUYIN_PARSE_DIR")
    parser.add_argument("--cookie", help="完整 cookie 字符串，或 sessionid=xxx")
    parser.add_argument("--cookie-file", help="从文件读取 cookie")
    parser.add_argument("--row", type=int, help="只处理指定行号")
    parser.add_argument("--asr-model", default="base",
                        choices=["tiny","base","small","medium","large"])
    parser.add_argument("--ai-method", default="skip",
                        choices=["skip","openai","deepseek","openai_api","deepseek_api"])
    parser.add_argument("--ai-key", help="API Key")
    parser.add_argument("--ai-base", help="OpenAI 兼容接口 base_url")
    parser.add_argument("--ai-model", help="模型名称")
    parser.add_argument("--interval", type=int, default=5,
                        help="每个视频间隔秒数")
    parser.add_argument("--update-index", action=argparse.BooleanOptionalAction, default=True,
                        help="处理完成后自动更新 video_index.json（默认开启，可用 --no-update-index 关闭）")
    args = parser.parse_args()

    config = {}
    config.update(load_env_file(str(Path(__file__).parent / "config/config.env")))
    config.update(load_env_file(args.config))

    # 读取 Cookie
    if args.cookie_file:
        cookie = open(args.cookie_file).read().strip()
    elif args.cookie:
        cookie = args.cookie
    elif config_get(config, "DOUYIN_SESSIONID"):
        sessionid = config_get(config, "DOUYIN_SESSIONID")
        cookie = sessionid if "sessionid=" in sessionid else f"sessionid={sessionid}"
    else:
        print("错误：需要 --cookie 或 --cookie-file")
        sys.exit(1)

    # 写入 cookie 文件（供 douyin_parse 使用）
    parser_dir = os.path.abspath(args.parser_dir)
    if not os.path.isdir(parser_dir):
        print(f"错误：douyin_parse 目录不存在: {parser_dir}")
        sys.exit(1)
    cookie_path = os.path.join(parser_dir, "douyin_cookie.txt")
    with open(cookie_path, "w") as f:
        f.write(cookie)
    print(f"Cookie 已写入 {cookie_path}")

    # 切换到 douyin_parse 目录，让 parser 能找到 cookie 文件
    os.chdir(parser_dir)
    parser_dy = get_douyin_parser(parser_dir)

    # AI 配置
    ai_method = normalize_ai_method(args.ai_method or config_get(config, "AI_METHOD", "skip"))
    default_api_key = ""
    if ai_method == "deepseek":
        default_api_key = config_get(config, "DEEPSEEK_API_KEY") or config_get(config, "AI_API_KEY")
        default_api_base = config_get(config, "DEEPSEEK_API_BASE", "https://api.deepseek.com/v1")
        default_model = config_get(config, "DEEPSEEK_MODEL", "deepseek-chat")
    else:
        default_api_key = config_get(config, "OPENAI_API_KEY") or config_get(config, "AI_API_KEY")
        default_api_base = config_get(config, "AI_API_BASE", "https://api.openai.com/v1")
        default_model = config_get(config, "AI_MODEL", "gpt-4o-mini")

    ai_config = {
        "method": ai_method,
        "api_key": args.ai_key or default_api_key,
        "api_base": args.ai_base or default_api_base,
        "model": args.ai_model or default_model,
    }

    # 打开 Excel
    if not os.path.exists(args.excel):
        print(f"Excel 文件不存在: {args.excel}")
        sys.exit(1)

    wb = openpyxl.load_workbook(args.excel)
    ws = wb[args.sheet] if args.sheet in wb.sheetnames else wb.active
    print(f"已打开 Excel：{args.excel}，共 {ws.max_row} 行")

    # 确定要处理的行
    if args.row:
        rows_to_process = [args.row]
    else:
        rows_to_process = [
            r for r in range(2, ws.max_row + 1)
            if (ws.cell(r, COL["状态"]).value or "未开始") in
               ("未开始", "失败")
            and (ws.cell(r, COL["链接"]).value or "").strip()
        ]

    if not rows_to_process:
        print("没有需要处理的链接（全部已完成或无链接）")
        if args.update_index:
            update_video_index(args.excel)
        sys.exit(0)

    print(f"共 {len(rows_to_process)} 条待处理")
    for idx, row in enumerate(rows_to_process):
        print(f"\n── 处理进度 {idx+1}/{len(rows_to_process)} ──")
        ok = process_row(ws, row, cookie, args.asr_model,
                         ai_config, parser_dy)
        # 保存进度（每条都保存）
        wb.save(args.excel)
        if idx < len(rows_to_process) - 1:
            print(f"等待 {args.interval} 秒...")
            time.sleep(args.interval)

    print(f"\n✅ 全部处理完成！Excel：{args.excel}")



    # ── 更新视频索引 ──────────────────────────────────────────────
    if args.update_index:
        update_video_index(args.excel)


def update_video_index(excel_path: str):
    """
    读取 Excel 中所有已完成（含 ASR 文案）的视频，
    生成描述并写入 video_index.json。
    追加模式：新视频追加，已有的（相同 sheet+row）则覆盖。
    """
    import re, json
    import sqlite3
    from pathlib import Path

    INDEX_PATH = Path(__file__).parent / "video_index.json"
    # 内容库 SQLite（与 web 端同一路径），用于关联整理稿状态
    DB_PATH = Path(__file__).parent / "output" / "video2text.db"

    # 整理稿聚合：aweme_id → 最新整理稿（类型/标题/时间），供索引条目标注「已整理」
    summary_by_aweme: dict[str, dict] = {}
    if DB_PATH.exists():
        try:
            conn = sqlite3.connect(str(DB_PATH))
            conn.row_factory = sqlite3.Row
            srows = conn.execute(
                "SELECT v.aweme_id, s.summary_type, s.title, s.updated_at "
                "FROM ai_summaries s JOIN videos v ON s.video_id = v.id "
                "WHERE s.status != 'failed' "
                "ORDER BY COALESCE(s.updated_at, '') DESC, s.id DESC"
            ).fetchall()
            conn.close()
            for s in srows:
                aweme = str(s["aweme_id"] or "")
                if aweme and aweme not in summary_by_aweme:
                    summary_by_aweme[aweme] = {
                        "summary_type": s["summary_type"] or "",
                        "summary_title": (s["title"] or "")[:60],
                        "summary_at": (s["updated_at"] or "")[:16],
                    }
        except Exception as e:
            print(f"   ⚠️ 读取内容库整理稿失败（跳过）：{e}")

    # 读取现有索引（用于去重，以视频ID为主key）
    existing = {}       # sheet:row → video dict
    existing_by_id = {}  # aweme_id → video dict（真正去重）
    if INDEX_PATH.exists():
        with open(INDEX_PATH, encoding="utf-8") as f:
            old = json.load(f)
        for v in old.get("videos", []):
            key = f"{v.get('sheet')}:{v.get('row')}"
            existing[key] = v
            vid = v.get("id", "") or ""
            if vid:
                existing_by_id[vid] = v

    wb = openpyxl.load_workbook(excel_path)
    updated_count = 0

    def header_map(ws) -> dict:
        aliases = {
            "抖音链接": "链接",
            "原始链接": "链接",
            "处理状态": "状态",
            "视频文案(ASR)": "视频文案ASR",
            "口播原文(ASR)": "口播原文",
            "关键词/摘要": "关键词摘要",
            "选题等级": "选题等级",
            "适合平台": "适合平台",
            "文章角度": "文章角度",
            "事实风险": "事实风险",
            "Word文档路径": "Word文档路径",
            "是否已发布": "是否已发布",
        }
        mapping = {}
        for c in range(1, ws.max_column + 1):
            raw = ws.cell(1, c).value
            if not raw:
                continue
            name = aliases.get(str(raw).strip(), str(raw).strip())
            mapping[name] = c
        return mapping

    def cell_value(ws, row: int, name: str, headers: dict, default=""):
        col = headers.get(name) or COL.get(name)
        if not col:
            return default
        value = ws.cell(row, col).value
        return default if value is None else value

    def infer_topic(title: str, tags: str, asr: str, author: str) -> str:
        combined = f"{title} {tags} {asr[:800]}"
        if "鹏宇" in author or any(k in combined for k in ["RAG", "提示词", "微调", "AI大模型"]):
            return "AI技术教程"
        if any(k in combined for k in ["流放", "POE", "异界", "BD", "开荒", "通货", "天赋"]):
            return "流放2攻略"
        if any(k in combined for k in ["暗黑", "暗黑4", "Diablo", "D4"]):
            return "暗黑4攻略"
        return "视频内容"

    def infer_article_score(title: str, tags: str, asr: str, status: str) -> str:
        combined = f"{title} {tags} {asr[:1200]}"
        if status == "已写文稿":
            return "已转文章"
        high_value = ["开荒", "保姆", "设置", "优化", "异界", "通货", "BD", "天赋", "攻略", "避坑"]
        risky = ["BUG", "bug", "刷出400", "版本答案", "最强", "必看"]
        score = sum(1 for k in high_value if k in combined)
        if score >= 3:
            return "A"
        if score >= 1:
            return "B-需核查" if any(k in combined for k in risky) else "B"
        return "C"

    def infer_fact_risk(title: str, tags: str, asr: str, note: str, article_score: str) -> str:
        combined = f"{title} {tags} {asr[:1200]}"
        risk_terms = ["BUG", "bug", "刷出400", "版本答案", "最强", "必看", "无限", "暴涨"]
        risks = []
        if any(k in combined for k in risk_terms):
            risks.append("标题或口播有强结论，写稿前核查机制/数值")
        if note:
            risks.append("Excel备注有异常")
        if "需核查" in article_score:
            risks.append("选题等级标记需核查")
        return "；".join(risks)

    for ws in wb.worksheets:
        headers = header_map(ws)
        for r in range(2, ws.max_row + 1):
            status = cell_value(ws, r, "状态", headers)
            if status not in ("已完成", "已写文稿"):
                continue
            asr = (str(cell_value(ws, r, "视频文案ASR", headers) or
                   cell_value(ws, r, "口播原文", headers) or "")).strip()
            if not asr:
                continue

            title   = str(cell_value(ws, r, "标题", headers) or "")
            author  = str(cell_value(ws, r, "作者", headers) or "")
            tags    = str(cell_value(ws, r, "标签", headers) or "")
            vid     = str(cell_value(ws, r, "视频ID", headers) or "")
            ctime   = str(cell_value(ws, r, "发布时间", headers) or "")
            source_url = str(cell_value(ws, r, "链接", headers) or "")
            video_url = str(cell_value(ws, r, "视频链接", headers) or "")
            cover_url = str(cell_value(ws, r, "封面URL", headers) or "")
            ai_copy = str(cell_value(ws, r, "AI优化文案", headers) or "")
            keywords = str(cell_value(ws, r, "关键词摘要", headers) or "")
            note = str(cell_value(ws, r, "备注", headers) or "")
            topic = infer_topic(title, tags, asr, author)
            article_score = str(cell_value(ws, r, "选题等级", headers) or infer_article_score(title, tags, asr, status))

            key = f"{ws.title}:{r}"
            # 旧条目（按视频 ID 匹配）：保留工作台回写的 published/performance，
            # Excel「是否已发布」列为空时不能把回写状态冲掉
            old = existing_by_id.get(vid) if vid else None
            existing[key] = {
                "id": str(vid),
                "sheet": ws.title,
                "row": r,
                "author": author,
                "title": title[:160],
                "tags": tags,
                "create_time": str(ctime)[:10] if ctime else "",
                "source_url": source_url,
                "video_url": video_url,
                "cover_url": cover_url,
                "status": status,
                "topic": topic,
                "article_score": article_score,
                "platform_suggestion": str(cell_value(ws, r, "适合平台", headers) or "待判断"),
                "article_angle": str(cell_value(ws, r, "文章角度", headers) or ""),
                "fact_risk": str(cell_value(ws, r, "事实风险", headers) or infer_fact_risk(title, tags, asr, note, article_score)),
                "word_doc_path": str(cell_value(ws, r, "Word文档路径", headers) or ""),
                "published": str(cell_value(ws, r, "是否已发布", headers) or ""),
                "keywords": keywords,
                "ai_copy_exists": bool(ai_copy.strip()),
                "transcript_length": len(asr),
                "transcript_snippet": asr[:800],
            }
            if old:
                if old.get("published") and not existing[key]["published"]:
                    existing[key]["published"] = str(old["published"])
                if old.get("performance"):
                    existing[key]["performance"] = old["performance"]
            # 整理稿状态（来自内容库 SQLite）：标注该视频已生成过哪种整理稿
            summary_info = summary_by_aweme.get(vid)
            if summary_info:
                existing[key]["summary"] = summary_info
            updated_count += 1

    # 启发式描述生成（LLM 提炼前先用规则顶一下）
    def _make_desc(v: dict) -> str:
        asr = v.get("transcript_snippet", "")[:500]
        cat = v.get("topic") or "视频内容"

        # 提取中文句子作为摘要
        sents = re.findall(r'[一-鿿][^。！？.\n]{4,50}[。！？.\n]?', asr)
        snippet = sents[0][:70].strip() if sents else asr[:70].strip()
        return f"【{cat}】{snippet}"

    videos = list(existing.values())
    # 按视频ID去重（保留最新的一行）
    seen_ids = set()
    unique_videos = []
    for v in videos:
        vid = v.get("id", "")
        if vid and vid in seen_ids:
            continue
        if vid:
            seen_ids.add(vid)
        unique_videos.append(v)
    videos = unique_videos
    for v in videos:
        v["description"] = _make_desc(v)

    new_index = {
        "name": "抖音视频内容库",
        "description": "抖音视频文案索引，用于从短视频素材快速筛选可二创文章的题材。",
        "version": "1.1",
        "updated": datetime.now().strftime("%Y-%m-%d %H:%M"),
        "total": len(videos),
        "schema": {
            "article_score": "A/B/C/已转文章。用于判断是否值得整理成文章，B-需核查表示可能有价值但必须补证据。",
            "topic": "粗分类，如流放2攻略、暗黑4攻略、AI技术教程。",
            "fact_risk": "事实风险提示。用于提醒写稿前必须联网核查或人工确认。",
            "word_doc_path": "已经转成文章后的 Word 文档路径，可由人工或后续脚本回填。",
            "summary": "最新整理稿状态（summary_type/summary_title/summary_at），来自内容库 ai_summaries。wechat_material=公众号素材已就绪。",
        },
        "videos": sorted(videos, key=lambda x: (x.get("sheet", ""), x.get("row", 0))),
    }

    # 备份上一代索引（video_index.json 事故教训：写坏/截断后 published/performance
    # 只有索引这一份，.bak 保留上一代给恢复留后路；备份失败不阻塞主写入）
    import shutil
    try:
        if INDEX_PATH.exists() and INDEX_PATH.stat().st_size > 0:
            shutil.copy2(str(INDEX_PATH), str(INDEX_PATH) + ".bak")
    except Exception as e:
        print(f"   ⚠️ 备份旧索引失败（继续写入）：{e}")

    # 原子写入：先写临时文件再替换，避免进程中断留下半个 JSON
    tmp_path = str(INDEX_PATH) + ".tmp"
    with open(tmp_path, "w", encoding="utf-8") as f:
        json.dump(new_index, f, ensure_ascii=False, indent=2)
    import os
    os.replace(tmp_path, str(INDEX_PATH))

    print(f"\n📋 索引已更新：{updated_count} 条视频 → {INDEX_PATH}")
    print(f"   共 {len(videos)} 条。描述为初版，后续对话中可让 AI 优化。")


if __name__ == "__main__":
    main()
