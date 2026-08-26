"""AI-friendly material workspace exports.

This module keeps the human-facing Excel file, SQLite content DB, JSONL files,
and local visual assets in sync enough for browsing and downstream generation.
"""

from __future__ import annotations

import json
import os
import re
import shutil
import sqlite3
import subprocess
import tempfile
from datetime import datetime
from pathlib import Path
from typing import Any

import openpyxl
import requests


ROOT = Path(__file__).resolve().parent.parent
OUTPUT_DIR = ROOT / "output"
DB_PATH = OUTPUT_DIR / "video2text.db"
AI_DATASET_DIR = OUTPUT_DIR / "ai_dataset"
RAW_DIR = OUTPUT_DIR / "raw" / "douyin"
ASSETS_DIR = OUTPUT_DIR / "assets"
COVERS_DIR = ASSETS_DIR / "covers"
KEYFRAMES_DIR = ASSETS_DIR / "keyframes"


HEADER_ALIASES = {
    "抖音链接": "链接",
    "原始链接": "链接",
    "处理状态": "状态",
    "视频文案(ASR)": "视频文案ASR",
    "口播原文(ASR)": "口播原文",
    "关键词/摘要": "关键词摘要",
}


BASE_COLUMNS = {
    "链接": 1,
    "状态": 2,
    "视频ID": 3,
    "作者": 4,
    "发布时间": 5,
    "标题": 6,
    "视频文案ASR": 7,
    "标签": 8,
    "封面URL": 9,
    "视频链接": 10,
    "口播原文": 11,
    "AI优化文案": 12,
    "AI备选标题": 13,
    "关键词摘要": 14,
    "备注": 15,
}


def ensure_dirs() -> None:
    for path in (AI_DATASET_DIR, RAW_DIR, COVERS_DIR, KEYFRAMES_DIR):
        path.mkdir(parents=True, exist_ok=True)


def export_workspace(
    excel_path: str | Path = OUTPUT_DIR / "抖音视频信息.xlsx",
    db_path: str | Path = DB_PATH,
    rows: tuple[int, int] | None = None,
) -> dict[str, Any]:
    """Export Excel/SQLite material into AI-friendly JSONL files.

    This is deliberately cheap: it does not download media. Use
    ``prepare_assets`` for cover/keyframe files.
    """
    ensure_dirs()
    records = load_excel_records(excel_path, rows=rows)
    write_jsonl(AI_DATASET_DIR / "videos.jsonl", [to_video_json(record) for record in records])
    write_jsonl(
        AI_DATASET_DIR / "transcripts.jsonl",
        [to_transcript_json(record) for record in records if record.get("transcript")],
    )
    stats = workspace_stats(records)
    (AI_DATASET_DIR / "workspace_manifest.json").write_text(
        json.dumps(
            {
                "updated_at": now(),
                "excel_path": str(Path(excel_path).resolve()),
                "db_path": str(Path(db_path).resolve()),
                **stats,
            },
            ensure_ascii=False,
            indent=2,
        ),
        encoding="utf-8",
    )
    return stats


def prepare_assets(
    excel_path: str | Path = OUTPUT_DIR / "抖音视频信息.xlsx",
    rows: tuple[int, int] | None = None,
    parser=None,
    max_keyframes_per_video: int = 4,
) -> dict[str, Any]:
    """Download covers and extract lightweight keyframes for available videos."""
    ensure_dirs()
    records = load_excel_records(excel_path, rows=rows)
    cover_ok = 0
    keyframe_ok = 0
    raw_ok = 0
    skipped = 0

    for record in records:
        aweme_id = record.get("aweme_id")
        if not aweme_id:
            skipped += 1
            continue

        raw = {}
        if parser is not None:
            try:
                raw = parser.parse_video(record.get("source_url") or record.get("video_url") or "")
            except Exception:
                raw = {}
        if raw:
            raw_ok += 1
            save_raw(aweme_id, raw)
            record = merge_parser_record(record, raw)

        cover_url = record.get("cover_url")
        if cover_url and download_cover(aweme_id, cover_url):
            cover_ok += 1

        nwm_url = record.get("nwm_url") or record.get("download_url")
        if nwm_url and extract_keyframes(aweme_id, nwm_url, max_frames=max_keyframes_per_video):
            keyframe_ok += 1

    export_workspace(excel_path, rows=rows)
    return {
        "records": len(records),
        "raw_saved": raw_ok,
        "covers": cover_ok,
        "keyframe_sets": keyframe_ok,
        "skipped": skipped,
    }


def list_workbench(db_path: str | Path = DB_PATH, query: str = "", author: str = "", status: str = "") -> list[dict[str, Any]]:
    conn = sqlite3.connect(str(db_path))
    conn.row_factory = sqlite3.Row
    try:
        where = []
        params: list[Any] = []
        if query:
            like = f"%{query}%"
            where.append("(title LIKE ? OR author LIKE ? OR transcript LIKE ? OR tags LIKE ?)")
            params.extend([like, like, like, like])
        if author:
            where.append("author = ?")
            params.append(author)
        if status:
            if status == "asr_missing":
                where.append("(transcript IS NULL OR transcript = '')")
            elif status == "asr_ready":
                where.append("(transcript IS NOT NULL AND transcript != '')")
            else:
                where.append("status = ?")
                params.append(status)
        clause = "WHERE " + " AND ".join(where) if where else ""
        rows = conn.execute(
            f"""
            SELECT *
            FROM videos
            {clause}
            ORDER BY updated_at DESC, source_sheet, source_row
            LIMIT 500
            """,
            params,
        ).fetchall()

        # 关联该视频的整理稿（含合并稿：source_video_ids 里包含该视频即算）
        summaries_by_video: dict[int, list[dict[str, Any]]] = {}
        for sr in conn.execute(
            "SELECT id, video_id, source_video_ids, summary_type, title FROM ai_summaries ORDER BY id"
        ).fetchall():
            try:
                vids = json.loads(sr["source_video_ids"] or "[]")
            except (TypeError, ValueError):
                vids = []
            if not vids and sr["video_id"]:
                vids = [sr["video_id"]]
            entry = {"id": sr["id"], "summary_type": sr["summary_type"], "title": sr["title"]}
            for vid in vids:
                summaries_by_video.setdefault(vid, []).append(entry)

        return [
            {**with_asset_status(dict(row)), "summaries": summaries_by_video.get(row["id"], [])}
            for row in rows
        ]
    finally:
        conn.close()


def load_excel_records(excel_path: str | Path, rows: tuple[int, int] | None = None) -> list[dict[str, Any]]:
    wb = openpyxl.load_workbook(str(excel_path), read_only=False, data_only=True)
    records: list[dict[str, Any]] = []
    try:
        for ws in wb.worksheets:
            headers = header_map(ws)
            start = rows[0] if rows else 2
            end = rows[1] if rows else ws.max_row
            for row in range(max(2, start), min(ws.max_row, end) + 1):
                source_url = cell(ws, row, "链接", headers)
                title = cell(ws, row, "标题", headers)
                transcript = cell(ws, row, "口播原文", headers) or cell(ws, row, "视频文案ASR", headers)
                if not source_url and not title and not transcript:
                    continue
                aweme_id = cell(ws, row, "视频ID", headers) or extract_aweme_id(source_url)
                records.append(
                    {
                        "platform": "douyin",
                        "source_sheet": ws.title,
                        "source_row": row,
                        "aweme_id": aweme_id,
                        "source_url": source_url,
                        "video_url": cell(ws, row, "视频链接", headers),
                        "cover_url": cell(ws, row, "封面URL", headers),
                        "author": cell(ws, row, "作者", headers),
                        "published_at": cell(ws, row, "发布时间", headers),
                        "title": title,
                        "status": cell(ws, row, "状态", headers),
                        "tags": split_tags(cell(ws, row, "标签", headers)),
                        "transcript": transcript,
                        "ai_copy": cell(ws, row, "AI优化文案", headers),
                        "keywords": split_tags(cell(ws, row, "关键词摘要", headers)),
                        "remark": cell(ws, row, "备注", headers),
                    }
                )
    finally:
        wb.close()
    return records


def header_map(ws) -> dict[str, int]:
    mapping = {}
    for col in range(1, ws.max_column + 1):
        raw = ws.cell(1, col).value
        if not raw:
            continue
        name = HEADER_ALIASES.get(str(raw).strip(), str(raw).strip())
        mapping[name] = col
    return mapping


def cell(ws, row: int, name: str, headers: dict[str, int]) -> str:
    col = headers.get(name) or BASE_COLUMNS.get(name)
    if not col:
        return ""
    value = ws.cell(row, col).value
    return "" if value is None else str(value)


def to_video_json(record: dict[str, Any]) -> dict[str, Any]:
    aweme_id = record.get("aweme_id") or ""
    return {
        **record,
        "content_direction": infer_content_direction(record),
        "suitable_platforms": ["小红书", "抖音", "快手", "头条"],
        "transcript_length": len(record.get("transcript") or ""),
        "cover_path": local_cover_path(aweme_id),
        "keyframe_paths": local_keyframe_paths(aweme_id),
        "updated_at": now(),
    }


def to_transcript_json(record: dict[str, Any]) -> dict[str, Any]:
    transcript = record.get("transcript") or ""
    return {
        "platform": "douyin",
        "aweme_id": record.get("aweme_id"),
        "source_sheet": record.get("source_sheet"),
        "source_row": record.get("source_row"),
        "title": record.get("title"),
        "author": record.get("author"),
        "transcript": transcript,
        "chunks": chunk_text(transcript),
    }


def workspace_stats(records: list[dict[str, Any]]) -> dict[str, Any]:
    return {
        "total": len(records),
        "with_asr": sum(1 for r in records if r.get("transcript")),
        "missing_asr": sum(1 for r in records if not r.get("transcript")),
        "with_cover_url": sum(1 for r in records if r.get("cover_url")),
    }


def write_jsonl(path: Path, items: list[dict[str, Any]]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", encoding="utf-8") as f:
        for item in items:
            f.write(json.dumps(item, ensure_ascii=False) + "\n")


def save_raw(aweme_id: str, raw: dict[str, Any]) -> None:
    path = RAW_DIR / f"{safe_id(aweme_id)}.json"
    path.write_text(json.dumps(raw, ensure_ascii=False, indent=2), encoding="utf-8")


def merge_parser_record(record: dict[str, Any], raw: dict[str, Any]) -> dict[str, Any]:
    merged = dict(record)
    desc = raw.get("desc") or ""
    merged["aweme_id"] = raw.get("aweme_id") or merged.get("aweme_id")
    merged["title"] = merged.get("title") or desc.split("#")[0].strip()
    merged["author"] = merged.get("author") or raw.get("author_nickname") or ""
    merged["cover_url"] = merged.get("cover_url") or raw.get("cover_url") or ""
    merged["nwm_url"] = raw.get("nwm_url") or ""
    merged["download_url"] = raw.get("nwm_url") or ""
    if not merged.get("tags"):
        merged["tags"] = split_tags(" ".join(f"#{p}" for p in desc.split("#")[1:]))
    return merged


def download_cover(aweme_id: str, url: str) -> bool:
    path = COVERS_DIR / f"{safe_id(aweme_id)}.jpg"
    if path.exists() and path.stat().st_size > 2048:
        return True
    try:
        resp = requests.get(url, headers=http_headers(), timeout=20)
        if resp.status_code == 200 and len(resp.content) > 1024:
            path.write_bytes(resp.content)
            return True
    except Exception:
        return False
    return False


def extract_keyframes(aweme_id: str, video_url: str, max_frames: int = 4) -> bool:
    target_dir = KEYFRAMES_DIR / safe_id(aweme_id)
    sheet = target_dir / "sheet.jpg"
    if sheet.exists() and sheet.stat().st_size > 4096:
        return True
    target_dir.mkdir(parents=True, exist_ok=True)
    tmpdir = Path(tempfile.mkdtemp(prefix="v2t_frames_"))
    video_path = tmpdir / "video.mp4"
    try:
        if not download_video(video_url, video_path):
            return False
        duration = probe_duration(video_path)
        if duration <= 0:
            return False
        times = frame_times(duration, max_frames)
        frames = []
        for idx, ts in enumerate(times, 1):
            out = target_dir / f"{idx:02d}.jpg"
            cmd = [
                "ffmpeg",
                "-y",
                "-ss",
                f"{ts:.2f}",
                "-i",
                str(video_path),
                "-frames:v",
                "1",
                "-q:v",
                "3",
                str(out),
            ]
            subprocess.run(cmd, capture_output=True, timeout=30)
            if out.exists() and out.stat().st_size > 2048:
                frames.append(out)
        if frames:
            make_contact_sheet(frames, sheet)
            return True
    except Exception:
        return False
    finally:
        shutil.rmtree(tmpdir, ignore_errors=True)
    return False


def download_video(url: str, path: Path) -> bool:
    try:
        subprocess.run(
            [
                "curl",
                "-L",
                "--max-time",
                "180",
                "-A",
                http_headers()["User-Agent"],
                "-H",
                "Referer: https://www.douyin.com/",
                "-o",
                str(path),
                url,
            ],
            capture_output=True,
            check=False,
            timeout=200,
        )
        return path.exists() and path.stat().st_size > 1024 * 50
    except Exception:
        return False


def probe_duration(video_path: Path) -> float:
    try:
        result = subprocess.run(
            [
                "ffprobe",
                "-v",
                "error",
                "-show_entries",
                "format=duration",
                "-of",
                "default=noprint_wrappers=1:nokey=1",
                str(video_path),
            ],
            capture_output=True,
            text=True,
            timeout=20,
        )
        return float((result.stdout or "0").strip() or 0)
    except Exception:
        return 0.0


def frame_times(duration: float, max_frames: int) -> list[float]:
    if max_frames <= 1:
        return [max(0.2, duration * 0.5)]
    start = max(0.3, duration * 0.12)
    end = max(start, duration * 0.88)
    step = (end - start) / (max_frames - 1)
    return [start + i * step for i in range(max_frames)]


def make_contact_sheet(frames: list[Path], out: Path) -> None:
    from PIL import Image, ImageDraw

    thumbs = []
    max_w, max_h = 360, 220
    for frame in frames:
        im = Image.open(frame).convert("RGB")
        im.thumbnail((max_w, max_h))
        canvas = Image.new("RGB", (max_w, max_h), (245, 245, 245))
        canvas.paste(im, ((max_w - im.width) // 2, (max_h - im.height) // 2))
        thumbs.append(canvas)
    width = max_w * len(thumbs)
    height = max_h + 30
    sheet = Image.new("RGB", (width, height), "white")
    draw = ImageDraw.Draw(sheet)
    for i, thumb in enumerate(thumbs):
        x = i * max_w
        sheet.paste(thumb, (x, 0))
        draw.text((x + 8, max_h + 8), f"{i + 1}", fill=(30, 30, 30))
    out.parent.mkdir(parents=True, exist_ok=True)
    sheet.save(out, quality=90)


def with_asset_status(row: dict[str, Any]) -> dict[str, Any]:
    aweme_id = row.get("aweme_id") or extract_aweme_id(row.get("source_url") or "")
    row["aweme_id"] = aweme_id
    row["transcript_length"] = len(row.get("transcript") or "")
    row["cover_path"] = local_cover_path(aweme_id)
    row["keyframe_paths"] = local_keyframe_paths(aweme_id)
    row["has_local_cover"] = bool(row["cover_path"])
    row["has_keyframes"] = bool(row["keyframe_paths"])
    row["content_direction"] = infer_content_direction(row)
    return row


def local_cover_path(aweme_id: str) -> str:
    if not aweme_id:
        return ""
    path = COVERS_DIR / f"{safe_id(aweme_id)}.jpg"
    return str(path.relative_to(OUTPUT_DIR)) if path.exists() else ""


def local_keyframe_paths(aweme_id: str) -> list[str]:
    if not aweme_id:
        return []
    directory = KEYFRAMES_DIR / safe_id(aweme_id)
    if not directory.exists():
        return []
    return [str(p.relative_to(OUTPUT_DIR)) for p in sorted(directory.glob("*.jpg"))]


def infer_content_direction(record: dict[str, Any]) -> str:
    text = " ".join(
        [
            str(record.get("title") or ""),
            " ".join(record.get("tags") or []) if isinstance(record.get("tags"), list) else str(record.get("tags") or ""),
            str(record.get("transcript") or "")[:800],
        ]
    )
    if any(k in text for k in ["Codex", "Agent", "提示词", "大模型", "RAG", "Claude", "Vibe"]):
        return "AI小白入门"
    return "待分类"


def chunk_text(text: str, chunk_size: int = 900) -> list[dict[str, Any]]:
    text = (text or "").strip()
    if not text:
        return []
    chunks = []
    for idx, start in enumerate(range(0, len(text), chunk_size), 1):
        chunk = text[start : start + chunk_size]
        chunks.append({"chunk_id": idx, "text": chunk, "length": len(chunk)})
    return chunks


def split_tags(value: str) -> list[str]:
    if not value:
        return []
    return [t.strip("# \t\r\n") for t in re.split(r"[,，\s]+", value) if t.strip("# \t\r\n")]


def extract_aweme_id(url: str) -> str:
    match = re.search(r"/video/(\d+)", str(url or ""))
    return match.group(1) if match else ""


def safe_id(value: str) -> str:
    return re.sub(r"[^0-9A-Za-z_-]+", "_", str(value or "unknown"))


def http_headers() -> dict[str, str]:
    return {
        "User-Agent": "Mozilla/5.0 (Macintosh; Intel Mac OS X 14_5) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/125.0 Safari/537.36",
        "Referer": "https://www.douyin.com/",
    }


def now() -> str:
    return datetime.now().strftime("%Y-%m-%d %H:%M:%S")
