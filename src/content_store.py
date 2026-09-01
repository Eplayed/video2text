"""SQLite content store for videos and AI-ready summaries."""

from __future__ import annotations

import json
import re
import sqlite3
from datetime import datetime
from pathlib import Path
from typing import Any

import openpyxl


BASE_COLUMNS = {
    "链接": 1,
    "状态": 2,
    "视频ID": 3,
    "作者": 4,
    "发布时间": 5,
    "标题": 6,
    "视频文案ASR": 7,
    "标签": 8,
    "封面URL": 9,
    "视频链接": 10,
    "口播原文": 11,
    "AI优化文案": 12,
    "AI备选标题": 13,
    "关键词摘要": 14,
    "备注": 15,
}


def connect(db_path: str | Path) -> sqlite3.Connection:
    conn = sqlite3.connect(str(db_path))
    conn.row_factory = sqlite3.Row
    conn.execute("PRAGMA foreign_keys = ON")
    init_db(conn)
    return conn


def init_db(conn: sqlite3.Connection) -> None:
    conn.executescript(
        """
        CREATE TABLE IF NOT EXISTS videos (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            aweme_id TEXT,
            source_sheet TEXT NOT NULL,
            source_row INTEGER NOT NULL,
            source_url TEXT,
            video_url TEXT,
            cover_url TEXT,
            author TEXT,
            title TEXT,
            status TEXT,
            published_at TEXT,
            tags TEXT,
            transcript TEXT,
            ai_copy TEXT,
            keywords TEXT,
            remark TEXT,
            updated_at TEXT NOT NULL,
            UNIQUE(source_sheet, source_row)
        );

        CREATE INDEX IF NOT EXISTS idx_videos_aweme_id ON videos(aweme_id);
        CREATE INDEX IF NOT EXISTS idx_videos_author ON videos(author);
        CREATE INDEX IF NOT EXISTS idx_videos_title ON videos(title);

        CREATE TABLE IF NOT EXISTS ai_summaries (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            video_id INTEGER NOT NULL REFERENCES videos(id) ON DELETE CASCADE,
            summary_type TEXT NOT NULL,
            title TEXT NOT NULL,
            outline TEXT,
            content TEXT,
            keywords TEXT,
            structured_data TEXT,
            model TEXT,
            status TEXT NOT NULL DEFAULT 'draft',
            source_video_ids TEXT,
            created_at TEXT NOT NULL,
            updated_at TEXT NOT NULL
        );

        CREATE INDEX IF NOT EXISTS idx_ai_summaries_type ON ai_summaries(summary_type);
        CREATE INDEX IF NOT EXISTS idx_ai_summaries_video_id ON ai_summaries(video_id);
        """
    )
    _ensure_column(conn, "ai_summaries", "source_video_ids", "TEXT")
    _ensure_column(conn, "ai_summaries", "structured_data", "TEXT")
    # 分类打标与 Dify 知识库同步字段（Excel 重同步不会覆盖这些列）
    _ensure_column(conn, "videos", "category", "TEXT")
    _ensure_column(conn, "videos", "ai_tags", "TEXT")
    _ensure_column(conn, "videos", "game", "TEXT")
    _ensure_column(conn, "videos", "dify_document_id", "TEXT")
    _ensure_column(conn, "videos", "dify_synced_at", "TEXT")
    conn.execute(
        """
        CREATE TABLE IF NOT EXISTS subscriptions (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            sec_uid TEXT UNIQUE NOT NULL,
            user_url TEXT NOT NULL,
            author TEXT DEFAULT '',
            category TEXT DEFAULT '',
            game TEXT DEFAULT '',
            auto_asr INTEGER DEFAULT 1,
            created_at TEXT NOT NULL,
            last_synced_at TEXT DEFAULT '',
            last_new_count INTEGER DEFAULT 0,
            last_sync_note TEXT DEFAULT ''
        )
        """
    )
    _ensure_column(conn, "subscriptions", "category", "TEXT DEFAULT ''")
    _ensure_column(conn, "subscriptions", "game", "TEXT DEFAULT ''")
    _ensure_column(conn, "subscriptions", "platform", "TEXT DEFAULT 'douyin'")
    conn.commit()


def sync_excel_to_db(excel_path: str | Path, db_path: str | Path) -> dict[str, int]:
    # The sync code performs many random cell lookups. In openpyxl read_only mode
    # each ws.cell access can re-parse worksheet XML, which is painfully slow for
    # ASR-heavy sheets. The workbook is small enough to load normally here.
    wb = openpyxl.load_workbook(str(excel_path), read_only=False, data_only=True)
    conn = connect(db_path)
    upserted = 0
    skipped = 0
    now = _now()

    try:
        for ws in wb.worksheets:
            headers = _header_map(ws)
            for row in range(2, ws.max_row + 1):
                source_url = _cell(ws, row, "链接", headers)
                title = _cell(ws, row, "标题", headers)
                transcript = (
                    _cell(ws, row, "口播原文", headers)
                    or _cell(ws, row, "视频文案ASR", headers)
                )
                if not source_url and not title and not transcript:
                    skipped += 1
                    continue

                conn.execute(
                    """
                    INSERT INTO videos (
                        aweme_id, source_sheet, source_row, source_url, video_url,
                        cover_url, author, title, status, published_at, tags,
                        transcript, ai_copy, keywords, remark, updated_at
                    )
                    VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                    ON CONFLICT(source_sheet, source_row) DO UPDATE SET
                        aweme_id=excluded.aweme_id,
                        source_url=excluded.source_url,
                        video_url=excluded.video_url,
                        cover_url=excluded.cover_url,
                        author=excluded.author,
                        title=excluded.title,
                        status=excluded.status,
                        published_at=excluded.published_at,
                        tags=excluded.tags,
                        transcript=excluded.transcript,
                        ai_copy=excluded.ai_copy,
                        keywords=excluded.keywords,
                        remark=excluded.remark,
                        updated_at=excluded.updated_at
                    """,
                    (
                        _cell(ws, row, "视频ID", headers),
                        ws.title,
                        row,
                        source_url,
                        _cell(ws, row, "视频链接", headers),
                        _cell(ws, row, "封面URL", headers),
                        _cell(ws, row, "作者", headers),
                        title,
                        _cell(ws, row, "状态", headers),
                        _cell(ws, row, "发布时间", headers),
                        _cell(ws, row, "标签", headers),
                        transcript,
                        _cell(ws, row, "AI优化文案", headers),
                        _cell(ws, row, "关键词摘要", headers),
                        _cell(ws, row, "备注", headers),
                        now,
                    ),
                )
                upserted += 1
        conn.commit()
    finally:
        conn.close()
        wb.close()

    return {"upserted": upserted, "skipped": skipped}


def list_summaries(
    db_path: str | Path,
    summary_type: str = "",
    query: str = "",
    limit: int = 200,
) -> list[dict[str, Any]]:
    conn = connect(db_path)
    try:
        where = []
        params: list[Any] = []
        if summary_type:
            where.append("s.summary_type = ?")
            params.append(summary_type)
        if query:
            like = f"%{query}%"
            where.append(
                "(s.title LIKE ? OR s.content LIKE ? OR v.title LIKE ? OR v.author LIKE ?)"
            )
            params.extend([like, like, like, like])
        clause = "WHERE " + " AND ".join(where) if where else ""
        rows = conn.execute(
            f"""
            SELECT
                s.id, s.summary_type, s.title, s.outline, s.content, s.keywords,
                s.structured_data,
                s.model, s.status, s.source_video_ids, s.created_at, s.updated_at,
                v.id AS video_id, v.aweme_id, v.author, v.title AS video_title,
                v.source_sheet, v.source_row, v.video_url
            FROM ai_summaries s
            JOIN videos v ON v.id = s.video_id
            {clause}
            ORDER BY s.updated_at DESC, s.id DESC
            LIMIT ?
            """,
            [*params, limit],
        ).fetchall()
        return [_dict(row) for row in rows]
    finally:
        conn.close()


def get_summary(db_path: str | Path, summary_id: int) -> dict[str, Any] | None:
    conn = connect(db_path)
    try:
        row = conn.execute(
            """
            SELECT
                s.*, v.aweme_id, v.author, v.title AS video_title,
                v.source_sheet, v.source_row, v.video_url, v.transcript
            FROM ai_summaries s
            JOIN videos v ON v.id = s.video_id
            WHERE s.id = ?
            """,
            (summary_id,),
        ).fetchone()
        return _dict(row) if row else None
    finally:
        conn.close()


def get_video_by_source(
    db_path: str | Path,
    sheet: str | None = None,
    row: int | None = None,
    video_id: int | None = None,
) -> dict[str, Any] | None:
    conn = connect(db_path)
    try:
        if video_id is not None:
            result = conn.execute("SELECT * FROM videos WHERE id = ?", (video_id,)).fetchone()
        else:
            result = conn.execute(
                "SELECT * FROM videos WHERE source_sheet = ? AND source_row = ?",
                (sheet, row),
            ).fetchone()
        return _dict(result) if result else None
    finally:
        conn.close()


def get_videos_by_ids(db_path: str | Path, video_ids: list[int]) -> list[dict[str, Any]]:
    if not video_ids:
        return []
    conn = connect(db_path)
    try:
        placeholders = ",".join("?" for _ in video_ids)
        rows = conn.execute(
            f"SELECT * FROM videos WHERE id IN ({placeholders})",
            video_ids,
        ).fetchall()
        by_id = {int(row["id"]): _dict(row) for row in rows}
        return [by_id[video_id] for video_id in video_ids if video_id in by_id]
    finally:
        conn.close()


def save_summary(
    db_path: str | Path,
    video_id: int,
    summary_type: str,
    result: dict[str, Any],
    model: str = "",
    status: str = "draft",
    source_video_ids: list[int] | None = None,
) -> int:
    conn = connect(db_path)
    now = _now()
    try:
        cursor = conn.execute(
            """
            INSERT INTO ai_summaries (
                video_id, summary_type, title, outline, content, keywords, structured_data,
                model, status, source_video_ids, created_at, updated_at
            )
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            """,
            (
                video_id,
                summary_type,
                result.get("title", ""),
                _json_or_text(result.get("outline", "")),
                result.get("content", ""),
                _json_or_text(result.get("keywords", "")),
                _json_or_text(_structured_payload(result)),
                model,
                status,
                _json_or_text(source_video_ids or [video_id]),
                now,
                now,
            ),
        )
        conn.commit()
        return int(cursor.lastrowid)
    finally:
        conn.close()


def update_summary(
    db_path: str | Path,
    summary_id: int,
    result: dict[str, Any],
    model: str = "",
    status: str = "draft",
    source_video_ids: list[int] | None = None,
) -> None:
    conn = connect(db_path)
    now = _now()
    try:
        conn.execute(
            """
            UPDATE ai_summaries SET
                title=?, outline=?, content=?, keywords=?, structured_data=?,
                model=?, status=?, source_video_ids=COALESCE(?, source_video_ids), updated_at=?
            WHERE id=?
            """,
            (
                result.get("title", ""),
                _json_or_text(result.get("outline", "")),
                result.get("content", ""),
                _json_or_text(result.get("keywords", "")),
                _json_or_text(_structured_payload(result)),
                model,
                status,
                _json_or_text(source_video_ids) if source_video_ids is not None else None,
                now,
                summary_id,
            ),
        )
        conn.commit()
    finally:
        conn.close()



def delete_summary(db_path: str | Path, summary_id: int) -> None:
    conn = connect(db_path)
    try:
        conn.execute("DELETE FROM ai_summaries WHERE id = ?", (summary_id,))
        conn.commit()
    finally:
        conn.close()


def delete_video(db_path: str | Path, sheet: str, row: int) -> bool:
    """删除 SQLite 中的视频及其直接关联的整理稿（合并稿保留，仅清理引用）。"""
    conn = connect(db_path)
    try:
        r = conn.execute(
            "SELECT id FROM videos WHERE source_sheet = ? AND source_row = ?", (sheet, row)
        ).fetchone()
        if not r:
            return False
        vid = r["id"]
        conn.execute("DELETE FROM ai_summaries WHERE video_id = ?", (vid,))
        # 合并稿 source_video_ids 里清理该视频引用
        for s in conn.execute(
            "SELECT id, source_video_ids FROM ai_summaries WHERE source_video_ids IS NOT NULL"
        ).fetchall():
            try:
                vids = json.loads(s["source_video_ids"] or "[]")
            except (TypeError, ValueError):
                continue
            if vid in vids:
                vids = [x for x in vids if x != vid]
                if not vids:
                    conn.execute("DELETE FROM ai_summaries WHERE id = ?", (s["id"],))
                else:
                    conn.execute(
                        "UPDATE ai_summaries SET source_video_ids = ? WHERE id = ?",
                        (json.dumps(vids), s["id"]),
                    )
        conn.execute("DELETE FROM videos WHERE id = ?", (vid,))
        conn.commit()
        return True
    finally:
        conn.close()


# ════════════════════════════════════════════════════════════════
# 订阅（作者主页增量同步）
# ════════════════════════════════════════════════════════════════

def add_subscription(
    db_path: str | Path,
    sec_uid: str,
    user_url: str,
    author: str = "",
    category: str = "",
    game: str = "",
    platform: str = "douyin",
) -> dict[str, Any]:
    """添加/更新订阅，返回订阅记录。"""
    conn = connect(db_path)
    try:
        conn.execute(
            """
            INSERT INTO subscriptions (sec_uid, user_url, author, category, game, platform, created_at)
            VALUES (?, ?, ?, ?, ?, ?, ?)
            ON CONFLICT(sec_uid) DO UPDATE SET
                user_url = excluded.user_url,
                author = CASE WHEN excluded.author != '' THEN excluded.author ELSE subscriptions.author END,
                category = CASE WHEN excluded.category != '' THEN excluded.category ELSE subscriptions.category END,
                game = CASE WHEN excluded.game != '' THEN excluded.game ELSE subscriptions.game END,
                platform = excluded.platform
            """,
            (sec_uid, user_url, author, category, game, platform, _now()),
        )
        conn.commit()
        row = conn.execute(
            "SELECT * FROM subscriptions WHERE sec_uid = ?", (sec_uid,)
        ).fetchone()
        return dict(row)
    finally:
        conn.close()


def update_subscription_author(db_path: str | Path, sub_id: int, author: str) -> bool:
    """修改订阅作者显示名。"""
    conn = connect(db_path)
    try:
        cur = conn.execute(
            "UPDATE subscriptions SET author = ? WHERE id = ?", (author.strip(), sub_id)
        )
        conn.commit()
        return cur.rowcount > 0
    finally:
        conn.close()


def update_subscription_category(
    db_path: str | Path,
    sub_id: int,
    category: str,
    game: str = "",
) -> bool:
    """修改订阅作者的分类/游戏。"""
    conn = connect(db_path)
    try:
        cur = conn.execute(
            "UPDATE subscriptions SET category = ?, game = CASE WHEN ? != '' THEN ? ELSE game END WHERE id = ?",
            (category, game, game, sub_id),
        )
        conn.commit()
        return cur.rowcount > 0
    finally:
        conn.close()


def list_subscriptions(db_path: str | Path) -> list[dict[str, Any]]:
    conn = connect(db_path)
    try:
        rows = conn.execute("SELECT * FROM subscriptions ORDER BY id").fetchall()
        return [dict(r) for r in rows]
    finally:
        conn.close()


def delete_subscription(db_path: str | Path, sub_id: int) -> bool:
    conn = connect(db_path)
    try:
        cur = conn.execute("DELETE FROM subscriptions WHERE id = ?", (sub_id,))
        conn.commit()
        return cur.rowcount > 0
    finally:
        conn.close()


def update_subscription_sync(
    db_path: str | Path,
    sub_id: int,
    new_count: int,
    note: str,
    author: str = "",
) -> None:
    conn = connect(db_path)
    try:
        conn.execute(
            """
            UPDATE subscriptions SET
                last_synced_at = ?, last_new_count = ?, last_sync_note = ?,
                author = CASE WHEN ? != '' THEN ? ELSE author END
            WHERE id = ?
            """,
            (_now(), new_count, note, author, author, sub_id),
        )
        conn.commit()
    finally:
        conn.close()


def insert_wechat_articles(
    db_path: str | Path,
    articles: list[dict[str, Any]],
    category: str = "",
    game: str = "",
) -> list[dict[str, Any]]:
    """将微信文章写入 videos 表，返回已插入的记录列表。

    每篇文章的 transcript 字段存正文 Markdown，跳过 ASR。
    用 aweme_id（URL hash）去重，source_sheet='微信文章'。
    """
    if not articles:
        return []
    conn = connect(db_path)
    inserted: list[dict[str, Any]] = []
    try:
        # 查已有 aweme_id（去重）
        existing = {
            r["aweme_id"]
            for r in conn.execute(
                "SELECT aweme_id FROM videos WHERE source_sheet = '微信文章' AND aweme_id IS NOT NULL"
            ).fetchall()
        }
        # 下一行号
        max_row = conn.execute(
            "SELECT COALESCE(MAX(source_row), 0) FROM videos WHERE source_sheet = '微信文章'"
        ).fetchone()[0]

        for idx, art in enumerate(articles):
            if art["aweme_id"] in existing:
                continue
            row_num = max_row + idx + 1
            g = game if category == "游戏攻略" else ""
            conn.execute(
                """
                INSERT INTO videos (
                    aweme_id, source_sheet, source_row, source_url,
                    author, title, status, published_at,
                    transcript, category, game, updated_at
                ) VALUES (?, '微信文章', ?, ?, ?, ?, '已完成', ?, ?, ?, ?, ?)
                """,
                (
                    art["aweme_id"], row_num, art["url"],
                    art["author"], art["title"],
                    art["published_at"], art["transcript"],
                    category, g, _now(),
                ),
            )
            inserted.append({**art, "source_sheet": "微信文章", "source_row": row_num})
        conn.commit()
    finally:
        conn.close()
    return inserted


def get_wechat_article_links(db_path: str | Path) -> set[str]:
    """返回已入库的微信文章 URL 集合（用于同步去重）。"""
    conn = connect(db_path)
    try:
        rows = conn.execute(
            "SELECT source_url FROM videos WHERE source_sheet = '微信文章' AND source_url IS NOT NULL"
        ).fetchall()
        return {r["source_url"] for r in rows}
    finally:
        conn.close()


def generate_summary(
    video: dict[str, Any],
    summary_type: str,
    ai_config: dict[str, str] | None = None,
) -> tuple[dict[str, Any], str, str]:
    ai_config = ai_config or {}
    method = (ai_config.get("method") or "skip").strip()
    api_key = ai_config.get("api_key", "")
    model = ai_config.get("model", "")
    if method != "skip" and api_key:
        ai_result = _generate_with_llm(video, summary_type, ai_config)
        if "error" not in ai_result:
            if summary_type == "ai_interview":
                ai_result = _normalize_interview_result(
                    ai_result,
                    fallback_title=f"AI面试题：{video.get('title') or '未命名视频'}",
                    source_lines=[f"{video.get('title') or '未命名视频'}（{video.get('author') or '未知作者'}）"],
                )
            return ai_result, model or method, "ai"
        fallback = _generate_local(video, summary_type)
        fallback["content"] += f"\n\n[AI生成失败，已保存规则版草稿：{ai_result['error']}]"
        return fallback, model or method, "draft"
    return _generate_local(video, summary_type), "local-template", "draft"


def generate_collection_summary(
    videos: list[dict[str, Any]],
    summary_type: str,
    ai_config: dict[str, str] | None = None,
) -> tuple[dict[str, Any], str, str]:
    ai_config = ai_config or {}
    method = (ai_config.get("method") or "skip").strip()
    api_key = ai_config.get("api_key", "")
    model = ai_config.get("model", "")
    if method != "skip" and api_key:
        ai_result = _generate_collection_with_llm(videos, summary_type, ai_config)
        if "error" not in ai_result:
            if summary_type == "ai_interview":
                ai_result = _normalize_interview_result(
                    ai_result,
                    fallback_title="AI面试题合集",
                    source_lines=[
                        f"{video.get('title') or '未命名视频'}（{video.get('author') or '未知作者'}）"
                        for video in videos
                    ],
                )
            return ai_result, model or method, "ai"
        fallback = _generate_collection_local(videos, summary_type)
        fallback["content"] += f"\n\n[AI生成失败，已保存规则版草稿：{ai_result['error']}]"
        return fallback, model or method, "draft"
    return _generate_collection_local(videos, summary_type), "local-template", "draft"


def _generate_with_llm(
    video: dict[str, Any],
    summary_type: str,
    ai_config: dict[str, str],
) -> dict[str, Any]:
    prompt = _prompt(video, summary_type)
    try:
        import openai

        client = openai.OpenAI(
            api_key=ai_config["api_key"],
            base_url=ai_config.get("api_base") or "https://api.openai.com/v1",
            timeout=120,
            max_retries=1,
        )
        resp = client.chat.completions.create(
            model=ai_config.get("model") or "gpt-4o-mini",
            messages=[{"role": "user", "content": prompt}],
            temperature=0.25,
            response_format={"type": "json_object"},
        )
        return _parse_llm_json(resp.choices[0].message.content)
    except Exception as exc:
        return {"error": str(exc)}


def _generate_collection_with_llm(
    videos: list[dict[str, Any]],
    summary_type: str,
    ai_config: dict[str, str],
) -> dict[str, Any]:
    prompt = _collection_prompt(videos, summary_type)
    try:
        import openai

        client = openai.OpenAI(
            api_key=ai_config["api_key"],
            base_url=ai_config.get("api_base") or "https://api.openai.com/v1",
            timeout=180,
            max_retries=1,
        )
        resp = client.chat.completions.create(
            model=ai_config.get("model") or "gpt-4o-mini",
            messages=[{"role": "user", "content": prompt}],
            temperature=0.25,
            response_format={"type": "json_object"},
        )
        return _parse_llm_json(resp.choices[0].message.content)
    except Exception as exc:
        return {"error": str(exc)}


def _parse_llm_json(text: Any) -> dict[str, Any]:
    """宽容解析 LLM 返回的 JSON：剥离 markdown 代码块、截取首个平衡的 {...}。"""
    text = str(text or "").strip()
    if not text:
        return {}
    # 剥离 ```json ... ``` 代码块包裹
    if text.startswith("```"):
        text = re.sub(r"^```[a-zA-Z]*\s*", "", text)
        text = re.sub(r"\s*```\s*$", "", text).strip()
    try:
        return json.loads(text)
    except json.JSONDecodeError:
        pass
    # 截取首个配平的花括号块（忽略字符串内的花括号）
    start = text.find("{")
    if start < 0:
        return {}
    depth = 0
    in_str = False
    escape = False
    for i, ch in enumerate(text[start:], start):
        if escape:
            escape = False
        elif ch == "\\":
            escape = True
        elif ch == '"':
            in_str = not in_str
        elif not in_str:
            if ch == "{":
                depth += 1
            elif ch == "}":
                depth -= 1
                if depth == 0:
                    try:
                        return json.loads(text[start : i + 1])
                    except json.JSONDecodeError:
                        return {}
    return {}


# AI 面试/技术类场景：公众号素材档案按面试题解析文风组织（区别于游戏攻略文风）
INTERVIEW_CATEGORIES = {"前端面试", "AI技术", "编程开发"}


def _is_interview_scene(video: dict[str, Any]) -> bool:
    """判断视频是否属于 AI 面试场景：AI面试工作表来源，或分类为技术/面试类。"""
    if (video.get("source_sheet") or "").strip() == "AI面试":
        return True
    return (video.get("category") or "").strip() in INTERVIEW_CATEGORIES


def _generate_local(video: dict[str, Any], summary_type: str) -> dict[str, Any]:
    title = video.get("title") or "未命名视频"
    transcript = (video.get("transcript") or "").strip()
    tags = video.get("tags") or ""
    author = video.get("author") or ""
    sentences = _sentences(transcript)
    key_points = sentences[:6] or [transcript[:160] or title]
    keywords = _keywords(f"{title} {tags} {transcript[:1000]}")

    if summary_type == "ai_interview":
        out_title = f"AI面试题：{title[:40]}"
        concepts = _interview_concepts(title, tags, transcript)
        if concepts:
            keywords = concepts
            key_points = [
                f"{concept}：{_evidence_around(transcript, concept)}"
                for concept in concepts
            ]
        questions = _build_interview_questions(
            key_points=key_points,
            keywords=keywords,
            prefix=f"v{video.get('id') or 'x'}",
            source_title=title,
        )
        outline = ["题目", "答案解析", "考察点", "追问", "来源视频"]
        content = _questions_to_markdown(
            title=out_title,
            questions=questions,
            source_lines=[f"{title}（{author}）"],
            note="规则版草稿基于 ASR 自动抽题，入库前建议人工核对术语、选项和答案。"
        )
    elif summary_type == "wechat_material":
        # 公众号素材档案（规则版草稿）：固定【标签】分节，下游 Dify 按节解析
        out_title = f"公众号素材：{title[:36]}"
        outline = ["【标题候选】", "【导语】", "【核心论点】", "【关键数据】", "【正文骨架】", "【金句摘录】", "【风险核查】"]
        data_lines = [s for s in sentences if re.search(r"\d+(\.\d+)?%|\d{2,}", s)][:5]
        if _is_interview_scene(video):
            # AI 面试场景：核心论点/正文骨架按面试题解析组织，并抽技术概念做关键词
            concepts = _interview_concepts(title, tags, transcript)
            if concepts:
                keywords = concepts
                arguments = [
                    f"{c}：{_evidence_around(transcript, c)[:50]}"
                    for c in concepts[:3]
                ]
            else:
                arguments = [s[:50] for s in sentences[:3]]
            skeleton = "## 问题背景\n## 核心原理\n## 工程实践\n## 考察点与追问"
        else:
            arguments = [s[:50] for s in sentences[:3]]
            skeleton = "## 事件背景\n## 关键细节\n## 玩家影响"
        content = (
            f"来源作者：{author}\n来源视频：{title}\n\n"
            "【标题候选】\n"
            f"1. {title[:40]}（配置 AI 后自动生成 3 个候选标题）\n\n"
            "【导语】\n"
            f"{(sentences[0] or title)[:40]}\n\n"
            "【核心论点】\n"
            + "\n".join(f"- {a}" for a in arguments)
            + "\n\n【关键数据】\n"
            + (("\n".join(f"- {s[:60]}" for s in data_lines) + "\n") if data_lines else "无\n")
            + "\n\n【正文骨架】\n" + skeleton + "\n\n"
            + "【金句摘录】\n"
            f"- {(sentences[1] if len(sentences) > 1 else '待补充')[:60]}（原话）\n\n"
            "【风险核查】\n"
            "- 规则版草稿未做事实核对，发布前核查所有数值/版本/时间线。\n\n"
            "> 规则版素材档案仅提供骨架；配置 AI 后重新生成可获得完整素材档案。"
        )
    elif summary_type == "guide_article":
        # 已废弃：成品文章由自媒体工作台生成，存量数据仅保留可读
        out_title = title[:40]
        outline = ["已废弃"]
        content = (
            f"来源作者：{author}\n\n"
            "> 该类型已废弃（成品文章由自媒体工作台生成）。存量草稿仅作参考，"
            "请改用「公众号文章素材」生成结构化素材档案。"
        )
    else:
        out_title = title[:40]
        outline = ["适用场景", "操作步骤", "关键机制", "避坑提醒"]
        content = (
            f"来源作者：{author}\n\n"
            "攻略要点：\n"
            + "\n".join(f"- {point}" for point in key_points)
            + "\n\n整理建议：把强结论、数值、版本时间重新核查后，再扩写成正式攻略。"
        )

    return {
        "title": out_title,
        "outline": outline,
        "content": content,
        "keywords": keywords,
        "questions": questions if summary_type == "ai_interview" else [],
        "structured_data": {"questions": questions} if summary_type == "ai_interview" else {},
    }


def _generate_collection_local(
    videos: list[dict[str, Any]],
    summary_type: str,
) -> dict[str, Any]:
    author_names = sorted({v.get("author") or "未知作者" for v in videos})
    corpus = "\n".join((v.get("transcript") or "")[:1200] for v in videos)
    keywords = _keywords(" ".join([v.get("title") or "" for v in videos]) + " " + corpus)
    source_lines = "\n".join(
        f"- {i}. {v.get('title') or '未命名视频'}（{v.get('author') or '未知作者'}）"
        for i, v in enumerate(videos, 1)
    )

    if summary_type == "ai_interview":
        title = f"AI面试题合集：{keywords[0] if keywords else '短视频整理'}"
        source_titles = [
            f"{video.get('title') or '未命名视频'}（{video.get('author') or '未知作者'}）"
            for video in videos
        ]
        key_points = []
        for video in videos:
            key_points.extend(_sentences(video.get("transcript") or "")[:2])
        concept_points = []
        concept_keywords = []
        for video in videos:
            concepts = _interview_concepts(
                video.get("title") or "",
                video.get("tags") or "",
                video.get("transcript") or "",
            )
            for concept in concepts:
                if concept not in concept_keywords:
                    concept_keywords.append(concept)
                    concept_points.append(f"{concept}：{_evidence_around(video.get('transcript') or '', concept)}")
        if concept_points:
            key_points = concept_points
            keywords = concept_keywords
        questions = _build_interview_questions(
            key_points=key_points,
            keywords=keywords,
            prefix=f"c{videos[0].get('id') if videos else 'x'}",
            source_title=title,
            max_questions=8,
        )
        outline = ["题目", "答案解析", "考察点", "追问", "来源视频"]
        content = _questions_to_markdown(
            title=title,
            questions=questions,
            source_lines=source_titles,
            note="合并题库草稿已按题目结构生成，入库前建议人工去重并校准答案。"
        )
    elif summary_type == "wechat_material":
        # 公众号素材档案（规则版草稿）：固定【标签】分节，下游 Dify 按节解析
        title = f"公众号素材：{keywords[0] if keywords else '短视频整理'}"
        outline = ["【标题候选】", "【导语】", "【核心论点】", "【关键数据】", "【正文骨架】", "【金句摘录】", "【风险核查】"]
        sentences = []
        for video in videos:
            sentences.extend(_sentences(video.get("transcript") or "")[:2])
        # 关键数据行：从口播里粗抽数字/百分比的句子
        data_lines = [s for s in sentences if re.search(r"\d+(\.\d+)?%|\d{2,}", s)][:5]
        if any(_is_interview_scene(v) for v in videos):
            # AI 面试场景：跨视频抽技术概念去重，论点/骨架按面试题解析组织
            concept_keywords = []
            arguments = []
            for video in videos:
                for concept in _interview_concepts(
                    video.get("title") or "", video.get("tags") or "", video.get("transcript") or ""
                ):
                    if concept not in concept_keywords and len(arguments) < 3:
                        concept_keywords.append(concept)
                        arguments.append(
                            f"{concept}：{_evidence_around(video.get('transcript') or '', concept)[:50]}"
                        )
            if not arguments:
                arguments = [s[:50] for s in sentences[:3]]
            if concept_keywords:
                keywords = concept_keywords + [k for k in keywords if k not in concept_keywords][:5]
            skeleton = "## 问题背景\n## 核心原理\n## 工程实践\n## 考察点与追问"
        else:
            arguments = [s[:50] for s in sentences[:3]]
            skeleton = "## 事件背景\n## 关键细节\n## 玩家影响"
        content = (
            f"来源视频：\n{source_lines}\n\n"
            "【标题候选】\n"
            f"1. {keywords[0] if keywords else '待定选题'}（配置 AI 后自动生成 3 个候选标题）\n\n"
            "【导语】\n"
            f"{sentences[0][:40] if sentences else '待补充'}\n\n"
            "【核心论点】\n"
            + "\n".join(f"- {a}" for a in arguments)
            + "\n\n【关键数据】\n"
            + (("\n".join(f"- {s[:60]}" for s in data_lines) + "\n") if data_lines else "无\n")
            + "\n\n【正文骨架】\n" + skeleton + "\n\n"
            + "【金句摘录】\n"
            f"- {sentences[1][:60] if len(sentences) > 1 else '待补充'}（原话）\n\n"
            "【风险核查】\n"
            "- 规则版草稿未做事实核对，发布前核查所有数值/版本/时间线。\n\n"
            "> 规则版素材档案仅提供骨架；配置 AI 后重新生成可获得完整素材档案。"
        )
    elif summary_type == "guide_article":
        # 已废弃：成品文章由自媒体工作台生成，存量数据仅保留可读
        title = f"（已废弃）攻略长文合集：{keywords[0] if keywords else '短视频整理'}"
        outline = ["已废弃"]
        key_points = []
        for video in videos:
            sentences = _sentences(video.get("transcript") or "")
            if sentences:
                key_points.append(sentences[0])
        content = (
            f"来源作者：{'、'.join(author_names)}\n\n"
            f"来源视频：\n{source_lines}\n\n"
            "> 该类型已废弃（成品文章由自媒体工作台生成）。存量草稿仅作参考，"
            "请改用「公众号文章素材」生成结构化素材档案。"
        )
    else:
        title = f"游戏攻略合集：{keywords[0] if keywords else '短视频整理'}"
        outline = ["攻略结论", "操作步骤", "机制解释", "避坑提醒", "来源视频"]
        key_points = []
        for video in videos:
            sentences = _sentences(video.get("transcript") or "")
            if sentences:
                key_points.append(sentences[0])
        content = (
            f"来源作者：{'、'.join(author_names)}\n\n"
            "来源视频：\n"
            f"{source_lines}\n\n"
            "攻略要点：\n"
            + "\n".join(f"- {point}" for point in key_points[:12])
            + "\n\n整理建议：合并稿适合做长文骨架，发布前核查版本、数值、装备名和机制。"
        )

    return {
        "title": title,
        "outline": outline,
        "content": content,
        "keywords": keywords,
        "questions": questions if summary_type == "ai_interview" else [],
        "structured_data": {"questions": questions} if summary_type == "ai_interview" else {},
    }


def _prompt(video: dict[str, Any], summary_type: str) -> str:
    if summary_type == "ai_interview":
        task = "整理成高质量 AI 面试题。必须生成 questions 数组，每道题是开放式面试题，包含题目、关键词答案、完整参考回答、答案解析、追问或工程实践延伸。"
        schema = """"questions": [
    {
      "id": "auto_q1",
      "type": "fill",
      "knowledgeTag": "知识点",
      "text": "题目文本",
      "answer": ["关键词答案"],
      "referenceAnswer": "完整参考回答，按问题背景、核心机制、工程边界、验证方法组织",
      "explain": "答案解析：说明这个题考什么，常见错误是什么",
      "deeper": "追问或工程实践延伸"
    }
  ]"""
    elif summary_type == "wechat_material":
        task = (
            "把口播整理成「公众号图文素材档案」——不是成品文章，是给下游 AI 写稿引擎消费的结构化素材。"
            "正文必须严格按以下【固定标签】分节，标签原样输出、顺序不变，方便下游按节解析：\n"
            "【标题候选】3 个公众号风格标题，15 字内，具体+数字+结果导向，不用惊叹号堆砌；\n"
            "【导语】一句话（30 字内）点出读者收益或痛点；\n"
            "【核心论点】最多 3 条，每条一句话，信息密度优先；\n"
            "【关键数据】把口播出现的数值/版本/日期/装备/技能名整理成 markdown 表格（表头：项目|数值|口径），口播没有的数据写“无”，严禁编造；\n"
            "【正文骨架】3-5 个 ## 小标题，按公众号阅读节奏组织（不写正文内容，只给骨架）；\n"
            "【金句摘录】直接引用口播原话 1-3 句，标注“（原话）”，保留口语感；\n"
            "【风险核查】发布前必须人工核对的强结论/数值清单，没有写“无强结论”。"
        )
        if video.get("game"):
            task += f"标题候选和正文骨架都要能看出游戏《{video['game']}》。"
        if _is_interview_scene(video):
            task += (
                "这是 AI 面试/技术类内容，素材档案按面试题解析文组织："
                "标题候选带「面试」「八股」「高频题」等场景词并点出具体技术点；"
                "【核心论点】按「问题→结论」组织，每条对应一个技术考点；"
                "【关键数据】优先整理技术栈/版本/性能指标/兼容性结论；"
                "【正文骨架】按「问题背景 → 核心原理 → 代码或工程实践 → 考察点与追问」组织；"
                "【风险核查】重点核对技术版本、性能数值、兼容性结论是否为口播原意。"
            )
        schema = """"outline": ["【标题候选】", "【导语】", "【核心论点】", "【关键数据】", "【正文骨架】", "【金句摘录】", "【风险核查】"],
  "content": "完整素材档案，按上述【固定标签】分节，markdown 表格用标准语法",
  "keywords": ["关键词1", "关键词2"]"""
    elif summary_type == "guide_article":
        # 已废弃：成品文章由自媒体工作台生成（API 层白名单已拦截，这里兜底）
        raise ValueError("guide_article 已废弃：成品文章由自媒体工作台生成，请使用 wechat_material 公众号素材档案")
    else:
        task = "整理成游戏攻略草稿，包含适用场景、步骤、机制解释、避坑提醒、事实核查清单。"
        schema = """"outline": ["一级要点1", "一级要点2"],
  "content": "完整正文，使用 Markdown",
  "keywords": ["关键词1", "关键词2"]"""
    return f"""你是内容整理助手。请只基于输入材料，不要编造事实，把短视频口播整理成结构化内容。

任务：{task}

输出 JSON：
{{
  "title": "整理后的标题",
  {schema},
  "keywords": ["关键词1", "关键词2"]
}}

视频标题：{video.get('title') or ''}
作者：{video.get('author') or ''}
游戏：{video.get('game') or ''}
标签：{video.get('tags') or ''}
口播 ASR：
{(video.get('transcript') or '')[:8000]}
"""


def _collection_prompt(videos: list[dict[str, Any]], summary_type: str) -> str:
    if summary_type == "ai_interview":
        task = "把多条视频合并整理成一套高质量 AI 面试题，去重相似内容。必须生成 questions 数组，每题是开放式面试题，包含题目、关键词答案、完整参考回答、答案解析、追问或工程实践延伸。"
        schema = """"questions": [
    {
      "id": "auto_q1",
      "type": "fill",
      "knowledgeTag": "知识点",
      "text": "题目文本",
      "answer": ["关键词答案"],
      "referenceAnswer": "完整参考回答，按问题背景、核心机制、工程边界、验证方法组织",
      "explain": "答案解析",
      "deeper": "追问或工程实践延伸"
    }
  ]"""
    elif summary_type == "wechat_material":
        task = (
            "把多条视频合并成一份「公众号图文素材档案」——不是成品文章，是给下游 AI 写稿引擎消费的结构化素材，去重相似内容。"
            "正文必须严格按以下【固定标签】分节，标签原样输出、顺序不变：\n"
            "【标题候选】3 个公众号风格标题，15 字内，具体+数字+结果导向；\n"
            "【导语】一句话（30 字内）点出读者收益；\n"
            "【核心论点】最多 3 条（跨视频去重合并）；\n"
            "【关键数据】markdown 表格（项目|数值|口径|来源视频编号），口播没有写“无”，严禁编造；\n"
            "【正文骨架】3-5 个 ## 小标题；\n"
            "【金句摘录】引用口播原话 1-3 句，标注“（原话）”；\n"
            "【风险核查】发布前必查的强结论/数值清单（标注来自哪条视频）。"
        )
        if any(_is_interview_scene(v) for v in videos):
            task += (
                "这批是 AI 面试/技术类内容，素材档案按面试题解析文组织："
                "标题候选带「面试」「八股」「高频题」等场景词并点出具体技术点；"
                "【核心论点】按「问题→结论」组织，每条对应一个技术考点；"
                "【关键数据】优先整理技术栈/版本/性能指标/兼容性结论；"
                "【正文骨架】按「问题背景 → 核心原理 → 代码或工程实践 → 考察点与追问」组织。"
            )
        schema = """"outline": ["【标题候选】", "【导语】", "【核心论点】", "【关键数据】", "【正文骨架】", "【金句摘录】", "【风险核查】"],
  "content": "完整素材档案，按【固定标签】分节，markdown 表格用标准语法",
  "keywords": ["关键词1", "关键词2"]"""
    elif summary_type == "guide_article":
        # 已废弃：成品文章由自媒体工作台生成（API 层白名单已拦截，这里兜底）
        raise ValueError("guide_article 已废弃：成品文章由自媒体工作台生成，请使用 wechat_material 公众号素材档案")
    else:
        task = "把多条视频合并整理成一篇游戏攻略合集，去重相似内容，输出适用场景、步骤、机制解释、避坑提醒和事实核查清单。"
        schema = """"outline": ["一级要点1", "一级要点2"],
  "content": "完整正文，使用 Markdown，并标出来源视频编号",
  "keywords": ["关键词1", "关键词2"]"""
    games = [v.get("game") for v in videos if v.get("game")]
    if games and summary_type != "ai_interview":
        unique_games = list(dict.fromkeys(games))
        game_note = "、".join(f"《{g}》" for g in unique_games)
        task += f"标题和正文点明游戏 {game_note}，方便同游戏读者检索。"
    blocks = []
    for i, video in enumerate(videos, 1):
        blocks.append(
            f"""[{i}]
标题：{video.get('title') or ''}
作者：{video.get('author') or ''}
游戏：{video.get('game') or ''}
标签：{video.get('tags') or ''}
口播 ASR：
{(video.get('transcript') or '')[:5000]}
"""
        )
    return f"""你是内容整理助手。请只基于输入材料，不要编造事实。

任务：{task}

输出 JSON：
{{
  "title": "整理后的合集标题",
  {schema},
  "keywords": ["关键词1", "关键词2"]
}}

输入视频：
{chr(10).join(blocks)}
"""


def _normalize_interview_result(
    result: dict[str, Any],
    fallback_title: str,
    source_lines: list[str],
) -> dict[str, Any]:
    questions = result.get("questions") or result.get("structured_data", {}).get("questions") or []
    normalized = []
    for idx, raw in enumerate(questions, 1):
        q_type = raw.get("type") or "fill"
        question = {
            "id": str(raw.get("id") or f"auto_q{idx}"),
            "type": q_type if q_type in {"single", "multi", "fill", "order", "locate", "interview"} else "fill",
            "knowledgeTag": str(raw.get("knowledgeTag") or raw.get("knowledge_tag") or "AI工程"),
            "text": str(raw.get("text") or raw.get("question") or ""),
            "answer": raw.get("answer") or [],
            "referenceAnswer": str(raw.get("referenceAnswer") or raw.get("reference_answer") or raw.get("参考答案") or ""),
            "explain": str(raw.get("explain") or raw.get("analysis") or raw.get("解析") or ""),
            "deeper": str(raw.get("deeper") or raw.get("follow_up") or raw.get("追问") or ""),
        }
        if question["type"] in {"single", "multi"}:
            question["options"] = _normalize_options(raw.get("options") or [])
        if question["text"]:
            normalized.append(question)

    if not normalized:
        normalized = _build_interview_questions(
            key_points=[result.get("content") or fallback_title],
            keywords=result.get("keywords") or [],
            prefix="auto",
            source_title=fallback_title,
        )

    title = result.get("title") or fallback_title
    return {
        "title": title,
        "outline": result.get("outline") or ["题目", "答案解析", "考察点", "追问", "来源视频"],
        "content": result.get("content") or _questions_to_markdown(
            title=title,
            questions=normalized,
            source_lines=source_lines,
            note="AI 已生成题库结构，入库前建议人工复核。",
        ),
        "keywords": result.get("keywords") or [q.get("knowledgeTag", "") for q in normalized[:8]],
        "questions": normalized,
        "structured_data": {"questions": normalized},
    }


def _build_interview_questions(
    key_points: list[str],
    keywords: list[str],
    prefix: str,
    source_title: str,
    max_questions: int = 5,
) -> list[dict[str, Any]]:
    points = [point.strip() for point in key_points if point and point.strip()]
    if not points:
        points = [source_title]
    tags = keywords or _keywords(" ".join(points))
    questions = []
    for idx, point in enumerate(points[:max_questions], 1):
        tag = tags[(idx - 1) % len(tags)] if tags else "AI工程"
        concise = _compact_text(point, 180)
        answer = _build_reference_answer(tag, concise)
        questions.append({
            "id": f"{prefix}q{idx}",
            "type": "fill",
            "knowledgeTag": tag,
            "text": f"请结合视频内容，系统解释「{tag}」解决了什么问题、核心方法是什么，以及它的边界或注意事项。",
            "answer": [tag],
            "referenceAnswer": answer,
            "explain": f"这题考察的不是名词记忆，而是能否把视频里的口播信息整理成「问题 -> 方法 -> 边界 -> 验证」的工程化回答。核心依据：{concise}",
            "deeper": f"如果你在真实项目中使用「{tag}」，你会设计哪些指标或测试来证明它确实有效？",
        })
    return questions


def _build_reference_answer(tag: str, evidence: str) -> str:
    return (
        f"可以从四层来回答。\n"
        f"1. 问题背景：材料围绕「{tag}」讨论的不是抽象定义，而是实际使用中会遇到的效果不稳定、理解偏差或落地困难。\n"
        f"2. 核心机制：结合视频依据，可以概括为：{evidence}\n"
        f"3. 工程边界：回答时要说明它适合解决什么问题，不适合解决什么问题，哪些前提条件缺失时容易失效。\n"
        f"4. 验证方法：可以通过对比实验、失败样例、输出质量、准确率、延迟、成本或人工复核结果来判断方案是否真的有效。"
    )


def _interview_concepts(title: str, tags: str, transcript: str) -> list[str]:
    text = f"{title} {tags} {transcript}"
    candidates = [
        ("提示词六步法", ["六步法", "提示词"]),
        ("角色设定", ["角色", "帽子", "专业模式"]),
        ("Context 背景", ["Context", "背景"]),
        ("Task 任务", ["Task", "任务", "明确的动词"]),
        ("输出格式约束", ["格式", "表格", "输出"]),
        ("RAG 检索增强生成", ["RAG", "检索增强生成", "外挂一个大脑"]),
        ("向量数据库", ["向量", "书局库", "数据库"]),
        ("切分与重叠", ["切分", "重叠", "窗口长度"]),
        ("幻觉问题", ["幻觉", "胡说八道", "瞎编"]),
        ("知识时效性", ["知识之后", "数据截止", "今天早上"]),
        ("私有数据接入", ["数据私有", "内网", "企业"]),
        ("RAG 与微调区别", ["微调", "开卷考试", "闭卷"]),
        ("Agent 工具调用", ["工具", "Action", "Observation"]),
        ("异步并发", ["async", "并发", "I/O"]),
        ("流式输出", ["流式", "NDJSON", "SSE"]),
    ]
    concepts = []
    for concept, needles in candidates:
        if any(needle.lower() in text.lower() for needle in needles):
            concepts.append(concept)
    if concepts:
        return concepts[:6]
    return _keywords(f"{title} {tags} {transcript[:600]}")[:5]


def _evidence_around(text: str, term: str, radius: int = 150) -> str:
    if not text:
        return term
    lower = text.lower()
    aliases = {
        "提示词六步法": ["六步法", "提示词", "下命令"],
        "角色设定": ["帽子", "专业模式", "绝色", "角色"],
        "Context 背景": ["Context", "背景", "老板", "汇报"],
        "Task 任务": ["Task", "任务", "出租车", "明确"],
        "输出格式约束": ["格式", "表格", "输出", "小座文"],
        "RAG 检索增强生成": ["RAG", "外挂一个大脑", "检索增强生成"],
        "向量数据库": ["向量", "书局库", "数据库"],
        "切分与重叠": ["切分", "重叠", "窗口长度"],
        "幻觉问题": ["幻觉", "胡说八道", "瞎编"],
        "知识时效性": ["知识之后", "数据截止", "今天早上"],
        "私有数据接入": ["数据私有", "内网", "企业"],
        "RAG 与微调区别": ["微调", "开卷考试", "闭卷"],
    }
    needles = [
        term.lower(),
        *[alias.lower() for alias in aliases.get(term, [])],
        *[part.lower() for part in re.split(r"[\s/与]+", term) if len(part) >= 2],
    ]
    positions = [lower.find(needle) for needle in needles if lower.find(needle) >= 0]
    pos = min(positions) if positions else 0
    start = max(0, pos - radius)
    end = min(len(text), pos + radius)
    return _compact_text(text[start:end], 220)


def _questions_to_markdown(
    title: str,
    questions: list[dict[str, Any]],
    source_lines: list[str],
    note: str,
) -> str:
    chunks = [f"# {title}", "", "## 来源视频"]
    chunks.extend(f"- {line}" for line in source_lines)
    chunks.append("")
    chunks.append("## 面试题")
    for idx, question in enumerate(questions, 1):
        chunks.extend([
            "",
            f"### Q{idx}. {question.get('text', '')}",
            "",
            f"- 考察点：{question.get('knowledgeTag', '')}",
        ])
        options = question.get("options") or []
        if options:
            chunks.append("- 选项：")
            for option in options:
                chunks.append(f"  - {option.get('value')}. {option.get('text')}")
        chunks.extend([
            f"- 关键词答案：{_answer_to_text(question.get('answer'))}",
            f"- 参考回答：{question.get('referenceAnswer') or question.get('reference_answer') or ''}",
            f"- 答案解析：{question.get('explain', '')}",
            f"- 追问：{question.get('deeper', '')}",
        ])
    chunks.extend(["", f"## 入库备注\n{note}"])
    return "\n".join(chunks)


def _normalize_options(options: list[Any]) -> list[dict[str, str]]:
    values = ["a", "b", "c", "d"]
    normalized = []
    for idx, option in enumerate(options[:4]):
        if isinstance(option, dict):
            text = str(option.get("text") or option.get("label") or "")
            value = str(option.get("value") or values[idx])
        else:
            text = str(option)
            value = values[idx]
        normalized.append({"text": text, "value": value})
    while len(normalized) < 4:
        idx = len(normalized)
        normalized.append({"text": f"干扰选项 {values[idx].upper()}", "value": values[idx]})
    return normalized


def _answer_to_text(answer: Any) -> str:
    if isinstance(answer, list):
        return "、".join(str(item) for item in answer)
    return str(answer)


def _header_map(ws) -> dict[str, int]:
    aliases = {
        "抖音链接": "链接",
        "原始链接": "链接",
        "处理状态": "状态",
        "视频文案(ASR)": "视频文案ASR",
        "口播原文(ASR)": "口播原文",
        "关键词/摘要": "关键词摘要",
    }
    mapping = {}
    for col in range(1, ws.max_column + 1):
        raw = ws.cell(1, col).value
        if not raw:
            continue
        name = aliases.get(str(raw).strip(), str(raw).strip())
        mapping[name] = col
    return mapping


def _cell(ws, row: int, name: str, headers: dict[str, int]) -> str:
    col = headers.get(name) or BASE_COLUMNS.get(name)
    if not col:
        return ""
    value = ws.cell(row, col).value
    return "" if value is None else str(value)


def _dict(row: sqlite3.Row) -> dict[str, Any]:
    return {key: row[key] for key in row.keys()}


def _ensure_column(conn: sqlite3.Connection, table: str, column: str, ddl: str) -> None:
    rows = conn.execute(f"PRAGMA table_info({table})").fetchall()
    if column not in {row["name"] for row in rows}:
        conn.execute(f"ALTER TABLE {table} ADD COLUMN {column} {ddl}")


def _json_or_text(value: Any) -> str:
    if isinstance(value, (list, dict)):
        return json.dumps(value, ensure_ascii=False)
    return "" if value is None else str(value)


def _structured_payload(result: dict[str, Any]) -> dict[str, Any]:
    if isinstance(result.get("structured_data"), dict):
        return result["structured_data"]
    if result.get("questions"):
        return {"questions": result["questions"]}
    # 攻略类（game_guide / guide_article）：按 ## 小标题切块，供程序化排版/抽取用
    sections = _split_markdown_sections(result.get("content") or "")
    if sections:
        return {
            "outline": result.get("outline") or [],
            "keywords": result.get("keywords") or [],
            "sections": sections,
        }
    return {}


def _split_markdown_sections(content: str) -> list[dict[str, str]]:
    """按 ## 小标题切 Markdown 正文；正文开头的引子归入 intro 段。"""
    sections: list[dict[str, str]] = []
    current_title = "intro"
    buf: list[str] = []
    for line in (content or "").splitlines():
        stripped = line.lstrip()
        if stripped.startswith("## "):
            if buf or sections:
                sections.append({"title": current_title, "body": "\n".join(buf).strip()})
            current_title = stripped[3:].strip()
            buf = []
        else:
            buf.append(line)
    if buf or sections:
        sections.append({"title": current_title, "body": "\n".join(buf).strip()})
    return [s for s in sections if s["title"] != "intro" or s["body"]]


def _sentences(text: str) -> list[str]:
    cleaned = re.sub(r"\s+", " ", text).strip()
    parts = re.split(r"[。！？!?；;\n]+", cleaned)
    sentences = [p.strip() for p in parts if len(p.strip()) >= 8]
    if len(sentences) <= 1 and len(cleaned) > 120:
        sentences = [
            cleaned[i:i + 110].strip()
            for i in range(0, min(len(cleaned), 660), 110)
            if cleaned[i:i + 110].strip()
        ]
    return sentences


def _compact_text(text: str, limit: int) -> str:
    value = re.sub(r"\s+", " ", str(text or "")).strip(" ，。,.")
    if len(value) <= limit:
        return value
    return value[:limit].rstrip(" ，。,.") + "..."


def _keywords(text: str) -> list[str]:
    words = re.findall(r"[A-Za-z0-9+#]{2,}|[\u4e00-\u9fff]{2,6}", text)
    stop = {"这个", "一个", "就是", "可以", "没有", "什么", "视频", "我们", "你们"}
    counts: dict[str, int] = {}
    for word in words:
        if word in stop:
            continue
        counts[word] = counts.get(word, 0) + 1
    return [word for word, _ in sorted(counts.items(), key=lambda item: -item[1])[:8]]


def _now() -> str:
    return datetime.now().strftime("%Y-%m-%d %H:%M:%S")


# ════════════════════════════════════════════════════════════════
# 分类打标
# ════════════════════════════════════════════════════════════════

CATEGORIES = [
    "前端面试", "AI技术", "游戏攻略", "编程开发",
    "产品设计", "数码评测", "汽车资讯", "商业财经",
    "职场成长", "生活日常", "其他",
]


def get_category_stats(db_path: str | Path) -> list[dict[str, Any]]:
    """分类 -> 数量统计（含未分类），供前端筛选 pills 渲染。"""
    conn = connect(db_path)
    try:
        rows = conn.execute(
            "SELECT COALESCE(NULLIF(category, ''), '未分类') AS cat, COUNT(*) AS n "
            "FROM videos GROUP BY cat ORDER BY n DESC"
        ).fetchall()
        return [{"category": r["cat"], "count": r["n"]} for r in rows]
    finally:
        conn.close()


# 渠道选题策略（与自媒体工作台 app.js 口径一致，集中管理便于后续调整）
# 头条号：7:2:1（攻略:资讯:试错），攻略特征词加权、资讯特征词降权、止损分类直接排除
# 公众号：知识沉淀渠道，AI/前端/编程等知识类标签优先，公众号素材稿权重更高
# 小红书：轻量话题渠道，近期热度权重放大，生活/攻略类标签优先
# 渠道策略版本：调整任何渠道的词表/分类/权重后 +1（media-workbench 兜底快照同步对齐）
CHANNEL_STRATEGY_VERSION = 2

RADAR_CHANNEL_STRATEGY: dict[str, dict[str, Any]] = {
    "toutiao": {
        "label": "头条号",
        "desc": "7:2:1 策略 · 攻略标签加权 · 资讯标签降权 · 止损分类与AI/面试类内容排除",
        "guide_kw": ["指南", "一览", "盘点", "怎么", "怎么选", "别乱", "别扔", "别急", "路线",
                      "攻略", "必看", "先看", "算清", "速通", "排行", "清单", "合集", "天价",
                      "底材", "掉落", "配装", "毕业", "流程", "对比", "避坑", "注意", "设置",
                      "帧数", "插件"],
        "news_kw": ["定档", "要开", "要来", "来了", "跳票", "更新", "补丁", "版本", "开服",
                     "上架", "发布", "免费", "上线", "解禁", "开启", "改了", "被砍", "加强",
                     "削弱", "预警", "风险", "紧急", "官宣", "泄露", "爆料", "开测", "停服",
                     "复刻"],
        # 历史CTR<1%止损分类 + AI/面试类分类（归公众号渠道，不进头条号选题）
        "excluded_categories": ["数码评测", "汽车资讯", "AI技术", "前端面试"],
        # AI/面试关键词：标题或标签命中即剔除（游戏类豁免，防误伤"AI代打""AI BD"等游戏语境）
        "exclude_kw": [
            "AI", "人工智能", "大模型", "智能体", "机器学习", "深度学习", "神经网络",
            "多模态", "微调", "提示词", "AIGC", "生成式", "幻觉",
            "llm", "gpt", "claude", "deepseek", "rag", "mcp", "agentic", "agent",
            "embedding", "diffusion", "webllm", "onnx", "codex", "dify",
            "langchain", "langgraph", "lang", "copilot", "midjourney", "vibe coding",
            "面试", "八股", "秋招", "春招", "简历", "求职", "leetcode", "刷题",
            "算法题", "offer",
            # 汽车/数码/硬件类：流量跑偏来源（2nm芯片/车展/路虎/NVIDIA驱动等阅读几乎为0），归数码渠道不进头条号
            "芯片", "2nm", "手机", "汽车", "车展", "路虎", "智驾", "预售",
            "nvidia", "显卡", "处理器", "小米18", "激光雷达",
        ],
        "exclude_kw_exempt_categories": ["游戏攻略", "生活日常"],
    },
    "wechat": {
        "label": "公众号",
        "desc": "知识沉淀优先 · AI/前端/编程类标签加权 · 公众号素材稿加权",
        # 粗分类（视频库 category）+ 细分类关键词（video_index 选题 topic，如「AI技术教程」），
        # 下游引擎用子串匹配打通两套口径
        "knowledge_categories": ["AI技术", "前端面试", "编程开发", "产品设计", "前端", "编程"],
        "summary_bonus_type": "wechat_material",
    },
    "xhs": {
        "label": "小红书",
        "desc": "近期话题优先 · 近期热度权重×3 · 生活/攻略类标签加权",
        # 同上：粗分类 + 细分类关键词（「流放2攻略」「暗黑4攻略」等）
        "light_categories": ["生活日常", "游戏攻略", "攻略", "生活"],
    },
}

# 全局策略阈值（所有渠道共享）：止损线、发布结构目标、结构提示统计窗口
CHANNEL_STRATEGY_GLOBAL: dict[str, Any] = {
    "lowCtrLine": 0.01,       # 止损线：同类已发≥2篇且均CTR低于该值的分类建议停更
    "mixTarget": [7, 2, 1],   # 头条号发布结构目标（攻略:资讯:其他）
    "mixRecentCount": 10,     # 结构提示统计的近期发布篇数
    "minStatCount": 2,        # 分类止损判定的最小样本数
}


def get_channel_strategy_config() -> dict[str, Any]:
    """渠道策略配置下发：video2text 是策略唯一权威源。

    自媒体工作台通过 /api/strategy/channels 拉取本配置驱动前端策略引擎，
    避免策略口径在两个项目各存一份拷贝（调整只改此处，全局生效）。
    """
    return {
        # 版本号：权威源调整策略口径时递增；media-workbench 兜底快照对齐此版本，
        # 不一致时前端控制台 warn（漂移可自动发现，不再依赖人肉同步）
        "version": CHANNEL_STRATEGY_VERSION,
        "global": dict(CHANNEL_STRATEGY_GLOBAL),
        "channels": {
            key: {k: (list(v) if isinstance(v, list) else v) for k, v in cfg.items()}
            for key, cfg in RADAR_CHANNEL_STRATEGY.items()
        },
    }


def _radar_kw_hit(text: str, kws: list[str]) -> bool:
    """关键词命中：中文/长英文词用子串；短词 "ai" 要求前后非英字母，
    兼容「AI开发」「本地化AI」等中英混排，同时防误伤 raid/chain 等。"""
    lower = (text or "").lower()
    if not lower:
        return False
    for kw in kws:
        k = kw.lower()
        if k == "ai":
            if re.search(r"(?<![a-z])ai(?![a-z])", lower):
                return True
        elif k and k in lower:
            return True
    return False


def get_topic_radar(
    db_path: str | Path,
    category: str = "",
    game: str = "",
    author: str = "",
    months: int = 6,
    channel: str = "",
) -> dict[str, Any]:
    """选题雷达：聚合某分类/游戏/作者下 ai_tags 的出现频次、近期热度、来源视频。

    channel 为空时不区分渠道（保持原行为）；否则按 RADAR_CHANNEL_STRATEGY
    调整聚合过滤与排序策略。

    返回：
      topics: 按热度排序的标签列表，每个含 count / recent_count / videos（前5条）
      total_videos: 该范围内视频总数
      window_days: 近期窗口
      channel_info: 当前渠道策略说明（channel 非空时）
    """
    strategy = RADAR_CHANNEL_STRATEGY.get(channel) or None
    conn = connect(db_path)
    try:
        where = ["COALESCE(ai_tags, '') != ''"]
        params: list[Any] = []
        if category:
            where.append("COALESCE(category, '') = ?")
            params.append(category)
        if game:
            where.append("COALESCE(game, '') = ?")
            params.append(game)
        if author:
            where.append("COALESCE(author, '') = ?")
            params.append(author)
        rows = conn.execute(
            f"SELECT id, ai_tags, title, author, published_at, source_sheet, source_row, category "
            f"FROM videos WHERE {' AND '.join(where)} "
            f"ORDER BY COALESCE(published_at, '') DESC, id DESC",
            params,
        ).fetchall()
        # 渠道策略：排除止损分类（如头条号历史CTR<1%的分类）
        if strategy and strategy.get("excluded_categories"):
            bad = set(strategy["excluded_categories"])
            rows = [r for r in rows if (r["category"] or "") not in bad]
        # 头条号策略：AI/面试类内容归公众号渠道，标题/标签命中关键词即剔除（游戏类豁免）
        if strategy and strategy.get("exclude_kw"):
            exempt = set(strategy.get("exclude_kw_exempt_categories") or [])
            kws = strategy["exclude_kw"]
            rows = [
                r for r in rows
                if (r["category"] or "") in exempt
                or not (
                    _radar_kw_hit(r["title"] or "", kws)
                    or _radar_kw_hit(r["ai_tags"] or "", kws)
                )
            ]
        # 整理稿感知：视频维度取最新一条 ai_summaries（类型/标题/时间），供选题雷达标注与排序
        summary_map: dict[int, dict[str, str]] = {}
        try:
            srows = conn.execute(
                "SELECT video_id, summary_type, title, updated_at FROM ai_summaries "
                "WHERE status != 'failed' "
                "ORDER BY COALESCE(updated_at, '') DESC, id DESC"
            ).fetchall()
            for s in srows:
                summary_map.setdefault(
                    int(s["video_id"]),
                    {
                        "summary_type": s["summary_type"] or "",
                        "summary_title": s["title"] or "",
                        "summary_at": (s["updated_at"] or "")[:16],
                    },
                )
        except Exception:
            summary_map = {}
        # 公众号渠道：已有公众号素材稿的视频单列计数（供加权）
        wechat_material_ids: set[int] = set()
        if strategy and strategy.get("summary_bonus_type"):  # noqa: SIM102
            try:
                mrows = conn.execute(
                    "SELECT DISTINCT video_id FROM ai_summaries "
                    "WHERE summary_type = ? AND status != 'failed'",
                    (strategy["summary_bonus_type"],),
                ).fetchall()
                wechat_material_ids = {int(m["video_id"]) for m in mrows}
            except Exception:
                wechat_material_ids = set()
    finally:
        conn.close()

    # 近期窗口（天）：取全部视频发布时间的中位跨度，或固定 90 天
    pub_days: list[float] = []
    for r in rows:
        t = (r["published_at"] or "").strip()
        if len(t) >= 10:
            try:
                d = datetime.strptime(t[:10], "%Y-%m-%d")
                pub_days.append((datetime.now() - d).days)
            except ValueError:
                pass
    window_days = 90
    if pub_days:
        span = max(pub_days) - min(pub_days)
        # 窗口取跨度的 1/3，最少 30 天，最多 180 天
        window_days = max(30, min(180, int(span / 3) if span > 0 else 90))

    # 噪声过滤：纯数字、游戏名及其去数字碎片（如「流放之路2」→ 滤掉「流放之路」「流放」「2」）
    noise: set[str] = set()
    if game:
        noise.add(game.lower())
        stripped = game
        while True:
            nxt = re.sub(r"[0-9]+$", "", stripped).strip()
            if not nxt or nxt == stripped:
                break
            noise.add(nxt.lower())
            stripped = nxt
        # 游戏名的连续中文前缀也滤掉（流放之路 → 流放）
        for i in range(2, len(game)):
            frag = game[:i]
            if frag and not re.search(r"[0-9a-zA-Z]", frag):
                noise.add(frag.lower())

    topic_map: dict[str, dict[str, Any]] = {}
    for r in rows:
        tags = [t.strip() for t in re.split(r"[;；,，/\s]+", r["ai_tags"] or "") if t.strip()]
        if not tags:
            continue
        is_recent = False
        t = (r["published_at"] or "").strip()
        if len(t) >= 10:
            try:
                d = datetime.strptime(t[:10], "%Y-%m-%d")
                is_recent = (datetime.now() - d).days <= window_days
            except ValueError:
                pass
        vid_summary = summary_map.get(int(r["id"])) if r["id"] is not None else None
        # 渠道策略：视频维度判定
        vcat = r["category"] or ""
        title_text = (r["title"] or "")
        v_strategy = ""  # ""中性 / "guide"攻略 / "news"资讯
        guide_kw = (strategy or {}).get("guide_kw") or []
        news_kw = (strategy or {}).get("news_kw") or []
        for kw in guide_kw:
            if kw in title_text:
                v_strategy = "guide"
                break
        if not v_strategy:
            for kw in news_kw:
                if kw in title_text:
                    v_strategy = "news"
                    break
        # 公众号知识类 / 小红书轻量类判定
        is_knowledge = bool(strategy) and vcat in ((strategy or {}).get("knowledge_categories") or [])
        is_light = bool(strategy) and vcat in ((strategy or {}).get("light_categories") or [])
        has_material = (r["id"] is not None and int(r["id"]) in wechat_material_ids) if strategy else False
        for tag in tags[:8]:  # 每条视频最多贡献 8 个标签，防止长尾噪声
            if tag.isdigit() or tag.lower() in noise or len(tag) < 2:
                continue
            key = tag.lower()  # 大小写合并（Token/token）
            entry = topic_map.setdefault(
                key,
                {"tag": tag, "count": 0, "recent_count": 0, "videos": []},
            )
            entry["count"] += 1
            if is_recent:
                entry["recent_count"] += 1
            # 渠道策略计数（供加权与前端徽标）
            if v_strategy == "guide":
                entry["guide_count"] = entry.get("guide_count", 0) + 1
            elif v_strategy == "news":
                entry["news_count"] = entry.get("news_count", 0) + 1
            if is_knowledge:
                entry["knowledge_count"] = entry.get("knowledge_count", 0) + 1
            if is_light:
                entry["light_count"] = entry.get("light_count", 0) + 1
            if has_material:
                entry["material_count"] = entry.get("material_count", 0) + 1
            if vid_summary:
                # 选题级整理稿聚合：数量 + 最新整理时间（不受 videos 前 5 条上限影响）
                entry["summary_count"] = entry.get("summary_count", 0) + 1
                if vid_summary["summary_at"] > entry.get("latest_summary_at", ""):
                    entry["latest_summary_at"] = vid_summary["summary_at"]
            if len(entry["videos"]) < 5:
                vinfo = {
                    "title": (r["title"] or "")[:40],
                    "author": r["author"] or "",
                    "published_at": (r["published_at"] or "")[:10],
                    "sheet": r["source_sheet"],
                    "row": r["source_row"],
                }
                if vid_summary:
                    vinfo.update(vid_summary)
                entry["videos"].append(vinfo)

    topics = sorted(topic_map.values(), key=lambda x: (-x["count"], -x["recent_count"]))
    # 热度分 = 总频次 + 近期频次×2（近期内容权重更高）
    for t in topics:
        t["heat"] = t["count"] + t["recent_count"] * 2
        # 选题内视频按最新整理稿时间倒序（已整理的排前面，方便直接取材）
        t["videos"].sort(key=lambda v: v.get("summary_at", ""), reverse=True)

    if not strategy:
        # 不限渠道：保持原排序（最新整理稿优先，热度垫底）
        topics.sort(
            key=lambda x: (x.get("latest_summary_at") or "", x["heat"]),
            reverse=True,
        )
        return {
            "topics": topics,
            "total_videos": len(rows),
            "window_days": window_days,
        }

    # ── 渠道策略排序 ──
    if channel == "toutiao":
        # 头条号 7:2:1：攻略标签加权、资讯标签降权，攻略型选题排前面
        for t in topics:
            t["heat"] = (
                t["count"]
                + t["recent_count"] * 2
                + t.get("guide_count", 0) * 2
                - t.get("news_count", 0)
            )
            gc, nc = t.get("guide_count", 0), t.get("news_count", 0)
            if gc > 0 and gc >= nc:
                t["strategy"] = "guide"
            elif nc > 0:
                t["strategy"] = "news"
            else:
                t["strategy"] = ""
        topics.sort(key=lambda x: (x["strategy"] == "guide", x["heat"]), reverse=True)
    elif channel == "wechat":
        # 公众号：知识类标签优先，已有公众号素材稿的选题加权
        for t in topics:
            t["heat"] = (
                t["count"]
                + t["recent_count"]
                + t.get("knowledge_count", 0) * 2
                + t.get("material_count", 0) * 3
            )
            if t.get("knowledge_count", 0) > 0:
                t["strategy"] = "knowledge"
            else:
                t["strategy"] = ""
        topics.sort(
            key=lambda x: (x["strategy"] == "knowledge", x.get("material_count", 0), x["heat"]),
            reverse=True,
        )
    elif channel == "xhs":
        # 小红书：近期话题优先（近期权重×3），生活/攻略类标签加权
        for t in topics:
            t["heat"] = (
                t["count"]
                + t["recent_count"] * 3
                + t.get("light_count", 0) * 2
            )
            if t.get("light_count", 0) > 0:
                t["strategy"] = "light"
            else:
                t["strategy"] = ""
        topics.sort(key=lambda x: (x.get("recent_count", 0), x["heat"]), reverse=True)

    return {
        "topics": topics,
        "total_videos": len(rows),
        "window_days": window_days,
        "channel_info": {
            "channel": channel,
            "label": strategy.get("label", ""),
            "desc": strategy.get("desc", ""),
        },
    }


def update_video_category(
    db_path: str | Path,
    video_id: int,
    category: str,
    ai_tags: str = "",
    game: str | None = None,
) -> bool:
    if category and category not in CATEGORIES:
        category = "其他"
    if category != "游戏攻略" and game is not None:
        game = ""
    conn = connect(db_path)
    try:
        cur = conn.execute(
            "UPDATE videos SET category = ?, "
            "ai_tags = CASE WHEN ? = '' THEN ai_tags ELSE ? END, "
            "game = CASE WHEN ? IS NULL THEN game ELSE ? END, "
            "updated_at = ? WHERE id = ?",
            (
                category,
                ai_tags.strip(), ai_tags.strip(),
                game, game,
                _now(), video_id,
            ),
        )
        conn.commit()
        return cur.rowcount > 0
    finally:
        conn.close()


def get_game_stats(db_path: str | Path) -> list[dict[str, Any]]:
    """按游戏名统计视频数（仅游戏攻略分类）。"""
    conn = connect(db_path)
    try:
        rows = conn.execute(
            "SELECT COALESCE(NULLIF(TRIM(game), ''), '未识别游戏') AS g, COUNT(*) AS n "
            "FROM videos WHERE category = '游戏攻略' GROUP BY g ORDER BY n DESC"
        ).fetchall()
        return [{"game": r["g"], "count": r["n"]} for r in rows]
    finally:
        conn.close()


def _classify_prompt(batch: list[dict[str, Any]]) -> str:
    items = []
    for v in batch:
        snippet = _compact_text(v.get("transcript") or "", 500)
        items.append(
            f'{{"key": "{v["sheet"]}::{v["row"]}", "title": "{(v.get("title") or "")[:60]}", '
            f'"author": "{(v.get("author") or "")[:20]}", "tags": "{(v.get("tags") or "")[:60]}", '
            f'"excerpt": "{snippet}"}}'
        )
    return (
        "你是内容分类助手。根据以下抖音视频的标题、作者、标签和口播摘要，"
        f'为每条视频从这些分类中选一个最贴切的：{"、".join(CATEGORIES)}。\n'
        "同时为每条视频提炼 3-5 个内容关键词（标签里没有的、视频实际讲的主题）。\n"
        "如果分类是游戏攻略，识别出视频讲的具体游戏名（如：流放之路2、原神、黑神话:悟空），"
        '填入 game 字段；其他分类 game 留空字符串。\n\n'
        "输入视频列表：\n" + "\n".join(items) + "\n\n"
        '输出 JSON：{"results": [{"key": "sheet::row", "category": "分类", "game": "游戏名或空", "tags": ["关键词", ...]}, ...]}\n'
        "要求：key 必须原样返回；只输出 JSON，不要输出其他内容。"
    )


def _classify_local(video: dict[str, Any]) -> dict[str, Any]:
    """无 AI 时的规则兜底分类。"""
    text = f'{video.get("title") or ""} {video.get("tags") or ""} {(video.get("transcript") or "")[:800]}'
    rules = [
        ("前端面试", r"面试|前端|js|javascript|css|vue|react|浏览器|闭包|promise"),
        ("游戏攻略", r"攻略|游戏|副本|装备|打法|阵容|通关|赛季"),
        ("AI技术", r"大模型|ai|llm|gpt|deepseek|提示词|agent|智能体|人工智能"),
        ("编程开发", r"代码|编程|开发|python|java|接口|数据库|算法"),
        ("职场成长", r"职场|副业|涨薪|简历|裁员|打工|牛马"),
    ]
    category = "其他"
    for cat, pattern in rules:
        if re.search(pattern, text, re.I):
            category = cat
            break
    return {"key": f'{video["sheet"]}::{video["row"]}', "category": category,
            "tags": _keywords(text)[:5]}


def classify_videos(
    db_path: str | Path,
    ai_config: dict[str, str],
    force: bool = False,
    progress_cb=None,
) -> dict[str, int]:
    """批量分类：优先 LLM，失败批次降级规则分类。progress_cb(done, total, note)。"""
    conn = connect(db_path)
    try:
        # 订阅同步入库的视频自带 category（继承订阅分类），但仍缺 ai_tags（选题雷达依赖）。
        # 因此只要 category 或 ai_tags 任一为空就补分类，而不是仅看 category。
        where = "" if force else (
            "WHERE COALESCE(category, '') = '' OR COALESCE(ai_tags, '') = ''"
        )
        rows = conn.execute(
            f"SELECT id, source_sheet, source_row, author, title, tags, transcript "
            f"FROM videos {where} ORDER BY id"
        ).fetchall()
        videos = []
        for r in rows:
            v = dict(r)
            v["sheet"] = v["source_sheet"]
            v["row"] = v["source_row"]
            videos.append(v)
    finally:
        conn.close()
    if not videos:
        return {"total": 0, "classified": 0, "llm_batches": 0}

    total = len(videos)
    done = 0
    use_llm = ai_config.get("method") != "skip" and ai_config.get("api_key")
    batch_size = 10
    results: list[dict[str, Any]] = []

    for i in range(0, total, batch_size):
        batch = videos[i : i + batch_size]
        if use_llm:
            data = _llm_chat(ai_config, _classify_prompt(batch), timeout=90)
            parsed = data.get("results") if isinstance(data, dict) else None
            if parsed:
                keyed = {str(v["sheet"]) + "::" + str(v["row"]): v for v in batch}
                for item in parsed:
                    key = str(item.get("key") or "")
                    if key in keyed:
                        cat = item.get("category") or "其他"
                        if cat not in CATEGORIES:
                            cat = "其他"
                        game = str(item.get("game") or "").strip()
                        if cat != "游戏攻略":
                            game = ""
                        tags = item.get("tags") or []
                        if isinstance(tags, list):
                            tags = " ".join(str(t) for t in tags[:5])
                        results.append(
                            {"id": keyed[key]["id"], "category": cat, "ai_tags": str(tags or ""),
                             "game": game}
                        )
                        continue
                # LLM 漏掉的条目用规则兜底
                got_keys = {r.get("key") for r in (parsed or [])}
                for v in batch:
                    if f'{v["source_sheet"]}::{v["source_row"]}' not in got_keys:
                        r = _classify_local(
                            {"sheet": v["source_sheet"], "row": v["source_row"],
                             "title": v["title"], "tags": v["tags"], "transcript": v["transcript"]}
                        )
                        results.append({"id": v["id"], "category": r["category"],
                                        "ai_tags": " ".join(r["tags"])})
            else:
                for v in batch:
                    r = _classify_local(
                        {"sheet": v["source_sheet"], "row": v["source_row"],
                         "title": v["title"], "tags": v["tags"], "transcript": v["transcript"]}
                    )
                    results.append({"id": v["id"], "category": r["category"],
                                    "ai_tags": " ".join(r["tags"])})
        else:
            for v in batch:
                r = _classify_local(
                    {"sheet": v["source_sheet"], "row": v["source_row"],
                     "title": v["title"], "tags": v["tags"], "transcript": v["transcript"]}
                )
                results.append({"id": v["id"], "category": r["category"],
                                "ai_tags": " ".join(r["tags"])})

        done += len(batch)
        if progress_cb:
            progress_cb(done, total, f"已分类 {done}/{total}")

    # 写回 DB
    conn = connect(db_path)
    try:
        for r in results:
            cat = r["category"] if r["category"] in CATEGORIES else "其他"
            game = r.get("game") or ""
            conn.execute(
                "UPDATE videos SET category = ?, ai_tags = ?, "
                "game = CASE WHEN ? = '' THEN game ELSE ? END, updated_at = ? WHERE id = ?",
                (cat, r["ai_tags"], game, game, _now(), r["id"]),
            )
        conn.commit()
    finally:
        conn.close()
    return {"total": total, "classified": len(results),
            "llm_batches": (total + batch_size - 1) // batch_size if use_llm else 0}


def _llm_chat(ai_config: dict[str, str], prompt: str, timeout: int = 90) -> dict[str, Any]:
    """通用 LLM JSON 调用（分类等场景复用），失败返回 {}。"""
    try:
        import openai

        client = openai.OpenAI(
            api_key=ai_config["api_key"],
            base_url=ai_config.get("api_base") or "https://api.openai.com/v1",
            timeout=timeout,
            max_retries=1,
        )
        resp = client.chat.completions.create(
            model=ai_config.get("model") or "gpt-4o-mini",
            messages=[{"role": "user", "content": prompt}],
            temperature=0.1,
            response_format={"type": "json_object"},
        )
        return _parse_llm_json(resp.choices[0].message.content)
    except Exception:
        return {}


# ════════════════════════════════════════════════════════════════
# Dify 知识库发布
# ════════════════════════════════════════════════════════════════

def get_dify_candidates(
    db_path: str | Path,
    keys: list[str] | None = None,
) -> list[dict[str, Any]]:
    """取待发布视频。keys 为 ["sheet::row", ...]；为空则取全部有转写的视频。"""
    conn = connect(db_path)
    try:
        rows = conn.execute(
            "SELECT v.id, v.source_sheet, v.source_row, v.author, v.title, v.tags, "
            "v.category, v.ai_tags, v.game, v.transcript, v.published_at, "
            "v.dify_document_id, v.cover_url, v.source_url "
            "FROM videos v WHERE COALESCE(v.transcript, '') != '' ORDER BY v.id"
        ).fetchall()
        items = [dict(r) for r in rows]
    finally:
        conn.close()
    if keys:
        keyset = set(keys)
        items = [v for v in items if f'{v["source_sheet"]}::{v["source_row"]}' in keyset]
    # 附带该视频的整理稿正文（丰富知识库内容）
    if items:
        conn = connect(db_path)
        try:
            for v in items:
                rows = conn.execute(
                    "SELECT title, content FROM ai_summaries WHERE video_id = ? AND content != ''",
                    (v["id"],),
                ).fetchall()
                v["summaries"] = [
                    {"title": r["title"], "content": (r["content"] or "")[:6000]} for r in rows
                ]
        finally:
            conn.close()
    return items


def build_dify_document(video: dict[str, Any]) -> tuple[str, str]:
    """构建 Dify 文档（名称, 正文）。"""
    title = (video.get("title") or "未命名视频").strip()
    name = f'[{video.get("author") or "未知"}] {title}'[:80]
    header_lines = [
        f"# {title}",
        "",
        f"- 作者：{video.get('author') or '未知'}",
        f"- 分类：{video.get('category') or '未分类'}",
        f"- 游戏：{video.get('game') or '无'}",
        f"- 标签：{video.get('ai_tags') or video.get('tags') or ''}",
        f"- 发布时间：{video.get('published_at') or '未知'}",
        f"- 来源：{video.get('source_url') or ''}",
        "",
        "## 口播原文",
        "",
        (video.get("transcript") or "").strip(),
    ]
    for s in video.get("summaries") or []:
        if s.get("content"):
            header_lines += ["", f"## 整理稿：{s['title']}", "", s["content"]]
    return name, "\n".join(header_lines)


def set_dify_doc(
    db_path: str | Path,
    video_id: int,
    document_id: str,
    synced_at: str,
) -> None:
    conn = connect(db_path)
    try:
        conn.execute(
            "UPDATE videos SET dify_document_id = ?, dify_synced_at = ?, updated_at = ? WHERE id = ?",
            (document_id, synced_at, _now(), video_id),
        )
        conn.commit()
    finally:
        conn.close()


def clear_dify_doc(db_path: str | Path, video_id: int) -> None:
    conn = connect(db_path)
    try:
        conn.execute(
            "UPDATE videos SET dify_document_id = NULL, dify_synced_at = NULL, updated_at = ? WHERE id = ?",
            (_now(), video_id),
        )
        conn.commit()
    finally:
        conn.close()
