"""
抖音用户主页视频批量采集
用法：
  python src/fetch_user_videos.py --url "https://www.douyin.com/user/MS4wLjAB..." --cookie "sessionid=xxx"
  python src/fetch_user_videos.py --video-url "https://v.douyin.com/xxx" --cookie "sessionid=xxx" --max-pages 5
"""
import argparse, sys, os, re, requests
from pathlib import Path
from urllib.parse import quote

ROOT = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(ROOT))
sys.path.insert(0, os.environ.get("DOUYIN_PARSE_DIR", "/tmp/douyin_parse"))

from douyin_video_parser import DouyinVideoParser


def resolve_short_url(short_url: str) -> str:
    """解析抖音短链接，返回真实 URL"""
    headers = {
        "User-Agent": "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36",
        "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8",
    }
    try:
        resp = requests.head(short_url, headers=headers, allow_redirects=True, timeout=10)
        return resp.url
    except Exception as e:
        print(f"解析短链接失败: {e}")
        return short_url


def extract_sec_uid_from_url(url: str) -> str:
    """从 URL 中提取 sec_uid"""
    patterns = [
        r"/user/([^/?\s]+)",
        r"sec_uid=([^&\s]+)",
    ]
    for pattern in patterns:
        m = re.search(pattern, url)
        if m:
            return m.group(1)
    return ""


def fetch_user_videos(
    url: str,
    cookie: str,
    max_pages: int = 10,
    count: int = 20,
    mode: str = "user_url",  # "user_url" or "video_url"
) -> dict:
    """
    获取用户主页视频列表。

    Returns:
        {
            "success": True,
            "user_url": "主页URL",
            "total": 30,
            "videos": [
                {
                    "aweme_id": "7644004659371281690",
                    "url": "https://www.douyin.com/video/7644004659371281690",
                    "share_url": "https://v.douyin.com/xxxxx/",  # 短链接待后续生成
                },
                ...
            ]
        }
    """
    # 确保 cookie 文件存在（DouyinVideoParser 依赖它）
    cookie_path = Path("/tmp/douyin_parse/douyin_cookie.txt")
    cookie_path.parent.mkdir(parents=True, exist_ok=True)
    with open(cookie_path, "w") as f:
        f.write(cookie)

    parser = DouyinVideoParser()
    parser.cookie = cookie

    # 处理短链接
    if "v.douyin.com" in url or "/share/" in url:
        print(f"🔗 解析短链接: {url}")
        url = resolve_short_url(url)
        print(f"   真实 URL: {url}")

    # 提取 sec_uid
    sec_uid = extract_sec_uid_from_url(url)
    if sec_uid:
        user_home = f"https://www.douyin.com/user/{sec_uid}"
        print(f"👤 用户主页: {user_home}")
        video_urls = parser.get_user_aweme_urls(user_home, max_pages=max_pages, count=count)
    elif mode == "video_url":
        # 从单个视频链接自动找到作者主页
        video_urls = parser.get_user_aweme_urls_from_video_url(
            url, max_pages=max_pages, count=count
        )
        # 尝试获取主页 URL
        user_home = parser.get_user_home_from_video_url(url)
    else:
        video_urls = parser.get_user_aweme_urls(
            url, max_pages=max_pages, count=count
        )
        user_home = url

    videos = []
    for v_url in video_urls:
        aweme_id = v_url.split("/video/")[-1].split("?")[0] if "/video/" in v_url else ""
        videos.append({
            "aweme_id": aweme_id,
            "url": v_url,
        })

    return {
        "success": True,
        "user_url": user_home or "",
        "total": len(videos),
        "videos": videos,
    }


def main():
    parser = argparse.ArgumentParser(description="抖音用户主页视频批量获取")
    parser.add_argument("--url", help="用户主页 URL（含 sec_uid）")
    parser.add_argument("--video-url", help="用户任意视频 URL（自动识别作者主页）")
    parser.add_argument("--cookie", required=True, help="Cookie（sessionid=xxx）")
    parser.add_argument("--max-pages", type=int, default=10, help="最大翻页数（默认10页，约200条）")
    parser.add_argument("--count", type=int, default=20, help="每页数量（默认20）")
    parser.add_argument("--output", help="输出到 Excel（可选，直接追加到 output/抖音视频信息.xlsx）")
    args = parser.parse_args()

    if not args.url and not args.video_url:
        print("❌ 需要提供 --url 或 --video-url")
        sys.exit(1)

    url = args.url or args.video_url
    mode = "user_url" if args.url else "video_url"

    print(f"🔍 正在获取用户视频列表...")
    result = fetch_user_videos(
        url=url,
        cookie=args.cookie,
        max_pages=args.max_pages,
        count=args.count,
        mode=mode,
    )

    if not result["success"]:
        print("❌ 获取失败")
        sys.exit(1)

    print(f"✅ 用户主页: {result['user_url']}")
    print(f"   共获取 {result['total']} 条视频")
    print()

    # 如果指定了 output，追加到 Excel
    if args.output:
        import openpyxl
        excel_path = Path(args.output)
        if not excel_path.exists():
            print(f"❌ Excel 文件不存在: {excel_path}")
            sys.exit(1)

        wb = openpyxl.load_workbook(str(excel_path))
        # 默认追加到第一个 sheet
        ws = wb.worksheets[0]
        next_row = ws.max_row + 1
        for r in range(2, ws.max_row + 1):
            if not ws.cell(r, 1).value:
                next_row = r
                break

        for i, v in enumerate(result["videos"]):
            row = next_row + i
            ws.cell(row, 1).value = v["url"]
            ws.cell(row, 2).value = "未开始"
            ws.cell(row, 3).value = v["aweme_id"]

        wb.save(str(excel_path))
        print(f"📝 已追加 {result['total']} 条链接到 Excel: {excel_path}")
        print(f"   起始行: {next_row}")
    else:
        # 输出到 stdout（JSON）
        import json
        print(json.dumps(result, ensure_ascii=False, indent=2))


if __name__ == "__main__":
    main()
